from doctest import master
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from openpyxl.styles.borders import Border, Side
import datetime
from PIL import ImageTk
from PIL import Image
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment
import pickle
from PIL import ImageGrab
from openpyxl.reader.excel import load_workbook
import os.path
import gc

gc.disable()

# main window
window = Tk()
window.title('Welcome to FAST Takeoff')
window.geometry('400x300+480+200')
style = ttk.Style()
style.theme_use('default')

# main window background
canvas = Canvas(window, height=300, width=500)
imagefile=PhotoImage(file='image/background.png')
image = canvas.create_image(0, 0, anchor='nw', image=imagefile)
canvas.pack(side='top')

# label user name and password
Label(window, text='Username:', bg="#395FB6", fg="white").place(x=99, y=150)
Label(window, text='Password:', bg="#395FB6", fg="white").place(x=99, y=190)
# input user name
var_usr_name = StringVar()
entry_usr_name = Entry(window, textvariable=var_usr_name)
entry_usr_name.place(x=160, y=150)
# input password
var_usr_pwd = StringVar()
entry_usr_pwd = Entry(window, textvariable=var_usr_pwd, show='*')
entry_usr_pwd.place(x=160, y=190)

# LOG IN DETAIL
def usr_log_in():
    # GET PASSWORD AND USER NAME
    usr_name = var_usr_name.get()
    usr_pwd = var_usr_pwd.get()
    # FIND INFORMATION IF NOT BUILD A NEW ONE
    try:
        with open('usr_info.pickle', 'rb') as usr_file:
            usrs_info = pickle.load(usr_file)
    except FileNotFoundError:
        with open('usr_info.pickle', 'wb') as usr_file:
            usrs_info = {'admin': 'admin'}
            pickle.dump(usrs_info, usr_file)
    # ENSURE USER NAME AND PASSWORD IS CORRECT
    if usr_name in usrs_info:
        if usr_pwd == usrs_info[usr_name]:
            messagebox.showinfo(title='WELCOME',
                                   message='Welcome:' + usr_name)
            window.destroy()

            def togglecheck(event):
                rowid = my_tree.identify_row(event.y)
                tag = my_tree.item(rowid, "tags")[0]
                tags = list(my_tree.item(rowid, "tags"))
                tags.remove(tag)
                my_tree.item(rowid, tags=tags)
                if tag == "checked":
                    my_tree.item(rowid, tags="unchecked")
                else:
                    my_tree.item(rowid, tags="checked")

            root = Tk()
            root.title('TAKING OFF LIST')
            root.geometry("760x600+300+50")

            # add some style
            style = ttk.Style()

            # pick a them
            style.theme_use('default')

            # treeview frame
            tree_frame = Frame(root)
            tree_frame.place(x=10, y=83)

            # scroll bar
            tree_scroll = Scrollbar(tree_frame)
            tree_scroll.pack(side=RIGHT, fill=Y)

            my_tree = ttk.Treeview(tree_frame, yscrollcommand=tree_scroll.set, selectmode="extended")
            my_tree.pack()

            # configure
            tree_scroll.configure(command=my_tree.yview_scroll)

            # define column
            my_tree['columns'] = ("ITEM", "CLAUSE", "UNIT", "QUANTITY")

            # Formate the column
            my_tree.column("#0", anchor=CENTER, width=50)
            my_tree.column("ITEM", anchor=W, width=220)
            my_tree.column("CLAUSE", anchor=CENTER, width=220)
            my_tree.column("UNIT", anchor=CENTER, width=120)
            my_tree.column("QUANTITY", anchor=CENTER, width=120)

            im_checked = ImageTk.PhotoImage(Image.open("checked.png"))
            im_unchecked = ImageTk.PhotoImage(Image.open("unchecked.png"))

            style = ttk.Style(my_tree)
            style.configure('Treeview', rowheight=28)

            # Create Headings
            my_tree.heading("#0", text="", anchor=CENTER)
            my_tree.heading("ITEM", text="ITEM", anchor=CENTER)
            my_tree.heading("CLAUSE", text="CLAUSE", anchor=CENTER)
            my_tree.heading("UNIT", text="UNIT", anchor=CENTER)
            my_tree.heading("QUANTITY", text="QUANTITY", anchor=CENTER)

            my_tree.tag_configure('checked', image=im_checked)
            my_tree.tag_configure('unchecked', image=im_unchecked)

            # add data
            data = []

            global count

            count = 0
            for record in data:
                my_tree.insert(parent='', index='end', lid=count, text="", values=(record))
                count += 1

            # pack to screen
            time_frame = Frame(root)
            time_frame.place(x=625, y=20)

            current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            time_label = Label(time_frame, text=current_time)
            time_label.pack()

            name_frame = Frame(root)
            name_frame.pack()

            project_frame = Frame(root)
            project_frame.place(x=10, y=55)

            taking_off_list_label = Label(name_frame, text="TAKING OFF LIST", font=(30))
            taking_off_list_label.pack(pady=10)

            add_frame = Frame(root)
            add_frame.place(x=10, y=380)

            item_label = Label(add_frame, text="ITEM")
            item_label.grid(row=0, column=0)

            clause_label = Label(add_frame, text="CLAUSE")
            clause_label.grid(row=0, column=1)

            unit_label = Label(add_frame, text="UNIT")
            unit_label.grid(row=0, column=2)

            quantity_label = Label(add_frame, text="QUANTITY")
            quantity_label.grid(row=0, column=3)

            project_name_label = Label(project_frame, text="PROJECT NAME")
            project_name_label.grid(row=0, column=0)

            taker_off_name_label = Label(project_frame, text="TAKER OFF")
            taker_off_name_label.grid(row=0, column=2)

            # ENTRY BOX

            project_box = Entry(project_frame, width=45)
            project_box.grid(row=0, column=1)

            taker_off_name_box = Entry(project_frame, width=28)
            taker_off_name_box.grid(row=0, column=3)

            item_box = Entry(add_frame, width=49)
            item_box.grid(row=1, column=0)

            # clause_box = Entry(add_frame, width=30, justify='center')
            clause = StringVar()
            clause_box = ttk.Combobox(add_frame, width=18, textvariable=clause, justify='center')
            clause_box.config(values=('D.12', 'D.25', 'F.3.5', 'F.3.7', 'F.3.8', 'F.3.15', 'F.3.16', 'F.8.1', 'F.8.4', 'F.10.1', 'G.3.1', 'G.3.3.a', 'L.4', 'L.17', 'S.2', 'S.11.1', 'S.1', 'U.3.1'))
            clause_box.grid(row=1, column=1)

            unit = StringVar()
            unit_box = ttk.Combobox(add_frame, width=18, textvariable=unit, justify='center')
            unit_box.config(values=('m', 'm2', 'm3', 'kg', 'item', 'no'))
            unit_box.grid(row=1, column=2)

            quantity_box = Entry(add_frame, width=20, justify='center')
            quantity_box.grid(row=1, column=3)
            quantity_box.insert(END, 0)

            # add item
            def add_item():
                if item_box.get() == "":
                    messagebox.showerror(message='PLEASE ENTER THE ITEM')
                elif clause_box.get() == "":
                    messagebox.showerror(message='PLEASE ENTER THE CLAUSE')
                elif unit_box.get() == "":
                    messagebox.showerror(message='PLEASE ENTER THE UNIT')

                else:
                    try:
                        input_quantity = float(quantity_box.get())
                    except ValueError:
                        messagebox.showerror(message="NUMERIC INPUT ONLY")
                        quantity_box.delete(0, END)

                    global count
                    my_tree.insert(parent='', index='end', iid=count,
                                   values=(item_box.get(), clause_box.get(), unit_box.get(), quantity_box.get()),
                                   tags="unchecked")
                    count += 1

                    # clear the box
                    item_box.delete(0, END)
                    clause_box.delete(0, END)
                    unit_box.delete(0, END)
                    quantity_box.delete(0, END)
                    quantity_box.insert(0, 0)


            #query list
            def query():
                root.withdraw()
                query = Tk()
                query.title('QUERY LIST')
                query.geometry("650x500+300+50")
                query_frame= Frame(query)

                # add query
                def add_query():
                    if query_item_box.get() == "":
                        messagebox.showerror(message='PLEASE ENTER THE ITEM')
                    elif query_assumption_box.get() == "":
                        messagebox.showerror(message='PLEASE ENTER THE ASSUMPTION')
                    else:
                        messagebox.showinfo("Query List", "Query updated.")

                    global count
                    my_tree_query.insert(parent='', index='end', iid=count,
                                   values=(query_item_box.get(), query_assumption_box.get()),
                                   tags="unchecked")
                    count += 1

                # delete query
                def remove_query():
                    x = my_tree_query.selection()[0]
                    my_tree_query.delete(x)

                def back_query():
                     query.withdraw()
                     root.deiconify()

                def togglecheck_query(event):
                    rowid = my_tree_query.identify_row(event.y)
                    tag = my_tree_query.item(rowid, "tags")[0]
                    tags = list(my_tree_query.item(rowid, "tags"))
                    tags.remove(tag)
                    my_tree_query.item(rowid, tags=tags)
                    if tag == "checked":
                        my_tree_query.item(rowid, tags="unchecked")
                    else:
                        my_tree_query.item(rowid, tags="checked")

                # button add query
                add_query_button = Button(query, text="ADD QUERY", command=add_query, width=15)
                add_query_button.place(x=100, y=450)

                # delete query
                delete_query = Button(query, text="DELETE QUERY", command=remove_query, width=15)
                delete_query.place(x=300, y=450)

                #back to taking off list
                back_query= Button(query,text="BACK", command=back_query, width=15)
                back_query.place(x=500, y=450)

                 # bind select record
                delete_query.bind("<ButtonRelease-1>", select_record)
                delete_query.bind("<Double-1>", togglecheck_query)

                # add some style
                style_query = ttk.Style()

                # pick a them
                style_query.theme_use('default')

                # treeview frame
                tree_query_frame = Frame(query)
                tree_query_frame.place(x=10, y=83)

                # scroll bar
                tree_query_scroll = Scrollbar(tree_query_frame)
                tree_query_scroll.pack(side=RIGHT, fill=Y)

                my_tree_query = ttk.Treeview(tree_query_frame, yscrollcommand=tree_query_scroll.set, selectmode="extended")
                my_tree_query.pack()

                # configure
                tree_query_scroll.configure(command=my_tree_query.yview_scroll)

                # define column
                my_tree_query['columns'] = ("ITEM", "ASSUMPTION")

                # Formate the column
                my_tree_query.column("#0", anchor=CENTER, width=50)
                my_tree_query.column("ITEM", anchor=CENTER, width=275)
                my_tree_query.column("ASSUMPTION", anchor=CENTER, width=275)

                query_im_checked = ImageTk.PhotoImage(Image.open("checked.png"))
                query_im_unchecked = ImageTk.PhotoImage(Image.open("unchecked.png"))

                style_query = ttk.Style(my_tree_query)
                style_query.configure('Treeview', rowheight=28)

                # Create Headings
                my_tree_query.heading("#0", text="", anchor=CENTER)
                my_tree_query.heading("ITEM", text="ITEM", anchor=CENTER)
                my_tree_query.heading("ASSUMPTION", text="ASSUMPTION", anchor=CENTER)

                my_tree_query.tag_configure('checked', image=query_im_checked)
                my_tree_query.tag_configure('unchecked', image=query_im_unchecked)

                # add data
                data = []

                global count

                count = 0
                for record in data:
                    my_tree_query.insert(parent='', index='end', lid=count, text="", values=(record))
                    count += 1

                # pack to screen
                query_time_frame = Frame(query)
                query_time_frame.place(x=525, y=20)

                query_current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                query_time_label = Label(query_time_frame, text=current_time)
                query_time_label.pack()

                query_name_frame = Frame(query)
                query_name_frame.pack()

                query_project_frame = Frame(query)
                query_project_frame.place(x=10, y=55)

                query_label = Label(query_name_frame, text="QUERY LIST", font=(30))
                query_label.pack(pady=10)

                query_add_frame = Frame(query)
                query_add_frame.place(x=10, y=380)

                query_item_label = Label(query_add_frame, text="ITEM")
                query_item_label.grid(row=0, column=1)

                query_assumption_label = Label(query_add_frame, text="ASSUMPTION")
                query_assumption_label.grid(row=0, column=2)

                query_project_name_label = Label(query_project_frame, text="PROJECT NAME")
                query_project_name_label.grid(row=0, column=0)

                query_taker_off_name_label = Label(query_project_frame, text="TAKER OFF")
                query_taker_off_name_label.grid(row=0, column=2)

                # ENTRY BOX

                query_project_box = Entry(query_project_frame, width=45)
                query_project_box.grid(row=0, column=1)

                query_taker_off_name_box = Entry(query_project_frame, width=28)
                query_taker_off_name_box.grid(row=0, column=3)

                query_item_box = Entry(query_add_frame, width=49)
                query_item_box.grid(row=1, column=1)

                query_assumption_box = Entry(query_add_frame, width=49, justify='center')
                query_assumption_box.grid(row=1, column=2)


            # taking off paper
            def taking_off_paper():
                root.withdraw()
                messagebox.showinfo("Attention!","Please confirm the unit before calculation.")


                def new_canvas(event):
                    canvas.delete('all')
                    display_pallete()


                def calculate():
                    r1c3.delete(0, END)
                    r2c3.delete(0, END)
                    r3c3.delete(0, END)
                    r4c3.delete(0, END)
                    r5c3.delete(0, END)
                    r6c3.delete(0, END)
                    r7c3.delete(0, END)
                    r8c3.delete(0, END)
                    r9c3.delete(0, END)
                    r10c3.delete(0, END)
                    r11c3.delete(0, END)
                    r12c3.delete(0, END)
                    r13c3.delete(0, END)
                    r14c3.delete(0, END)
                    r15c3.delete(0, END)
                    r16c3.delete(0, END)
                    r17c3.delete(0, END)
                    r18c3.delete(0, END)
                    r19c3.delete(0, END)
                    r20c3.delete(0, END)
                    r21c3.delete(0, END)
                    r22c3.delete(0, END)
                    r23c3.delete(0, END)
                    r24c3.delete(0, END)
                    r25c3.delete(0, END)
                    quantity_kg_o.delete(0, END)
                    quantity_paper_box.delete(0, END)
                    total_cost_box.delete(0, END)

                    if unit_box_cb.get() == 'm' or unit_box_cb.get() == 'no':

                        try:
                            input_quantity = float(r1c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r1c1.delete(0, END)
                            r1c1.insert(0, 0)

                        try:
                            input_quantity = float(r3c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r3c1.delete(0, END)
                            r3c1.insert(0, 0)

                        try:
                            input_quantity = float(r5c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r5c1.delete(0, END)
                            r5c1.insert(0, 0)

                        try:
                            input_quantity = float(r7c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r7c1.delete(0, END)
                            r7c1.insert(0, 0)

                        try:
                            input_quantity = float(r9c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r9c1.delete(0, END)
                            r9c1.insert(0, 0)

                        try:
                            input_quantity = float(r11c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r11c1.delete(0, END)
                            r11c1.insert(0, 0)

                        try:
                            input_quantity = float(r13c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r13c1.delete(0, END)
                            r13c1.insert(0, 0)

                        try:
                            input_quantity = float(r15c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r15c1.delete(0, END)
                            r15c1.insert(0, 0)

                        try:
                            input_quantity = float(r17c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r17c1.delete(0, END)
                            r17c1.insert(0, 0)

                        try:
                            input_quantity = float(r19c1.get())
                        except ValueError:
                            messagebox.showerror(message="INUMERIC INPUT ONLY")
                            r19c1.delete(0, END)
                            r19c1.insert(0, 0)

                        try:
                            input_quantity = float(r21c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r21c1.delete(0, END)
                            r21c1.insert(0, 0)

                        try:
                            input_quantity = float(r23c1.get())
                        except ValueError:
                            messagebox.showerror(message="INUMERIC INPUT ONLY")
                            r23c1.delete(0, END)
                            r23c1.insert(0, 0)

                        try:
                            input_quantity = float(r1c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r1c2.delete(0, END)
                            r1c2.insert(0, 0)

                        try:
                            input_quantity = float(r3c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r3c2.delete(0, END)
                            r3c2.insert(0, 0)

                        try:
                            input_quantity = float(r5c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r5c2.delete(0, END)
                            r5c2.insert(0, 0)

                        try:
                            input_quantity = float(r7c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r7c2.delete(0, END)
                            r7c2.insert(0, 0)

                        try:
                            input_quantity = float(r9c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r9c2.delete(0, END)
                            r9c2.insert(0, 0)

                        try:
                            input_quantity = float(r11c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r11c2.delete(0, END)
                            r11c2.insert(0, 0)

                        try:
                            input_quantity = float(r13c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r13c2.delete(0, END)
                            r13c2.insert(0, 0)

                        try:
                            input_quantity = float(r15c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r15c2.delete(0, END)
                            r15c2.insert(0, 0)

                        try:
                            input_quantity = float(r17c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r17c2.delete(0, END)
                            r17c2.insert(0, 0)

                        try:
                            input_quantity = float(r19c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r19c2.delete(0, END)
                            r19c2.insert(0, 0)

                        try:
                            input_quantity = float(r21c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r21c2.delete(0, END)
                            r21c2.insert(0, 0)

                        try:
                            input_quantity = float(r23c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r23c2.delete(0, END)
                            r23c2.insert(0, 0)

                        try:
                            input_quantity = float(rate_box.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            rate_box.delete(0, END)
                            rate_box.insert(0, 0)

                        c1r1 = float(r1c1.get())
                        c1r3 = float(r3c1.get())
                        c1r5 = float(r5c1.get())
                        c1r7 = float(r7c1.get())
                        c1r9 = float(r9c1.get())
                        c1r11 = float(r11c1.get())
                        c1r13 = float(r13c1.get())
                        c1r15 = float(r15c1.get())
                        c1r17 = float(r17c1.get())
                        c1r19 = float(r19c1.get())
                        c1r21 = float(r21c1.get())
                        c1r23 = float(r23c1.get())
                        c2r1 = float(r1c2.get())
                        c2r3 = float(r3c2.get())
                        c2r5 = float(r5c2.get())
                        c2r7 = float(r7c2.get())
                        c2r9 = float(r9c2.get())
                        c2r11 = float(r11c2.get())
                        c2r13 = float(r13c2.get())
                        c2r15 = float(r15c2.get())
                        c2r17 = float(r17c2.get())
                        c2r19 = float(r19c2.get())
                        c2r21 = float(r21c2.get())
                        c2r23 = float(r23c2.get())

                        r2c3.insert(0, '{:.2f}'.format(c1r1 * c2r1))
                        r4c3.insert(0, '{:.2f}'.format(c1r3 * c2r3))
                        r6c3.insert(0, '{:.2f}'.format(c1r5 * c2r5))
                        r8c3.insert(0, '{:.2f}'.format(c1r7 * c2r7))
                        r10c3.insert(0, '{:.2f}'.format(c1r9 * c2r9))
                        r12c3.insert(0, '{:.2f}'.format(c1r11 * c2r11))
                        r14c3.insert(0, '{:.2f}'.format(c1r13 * c2r13))
                        r16c3.insert(0, '{:.2f}'.format(c1r15 * c2r15))
                        r18c3.insert(0, '{:.2f}'.format(c1r17 * c2r17))
                        r20c3.insert(0, '{:.2f}'.format(c1r19 * c2r19))
                        r22c3.insert(0, '{:.2f}'.format(c1r21 * c2r21))
                        r24c3.insert(0, '{:.2f}'.format(c1r23 * c2r23))
                        res = float(r2c3.get()) + float(r4c3.get()) + float(r6c3.get()) + float(r8c3.get()) + float(
                            r10c3.get()) + float(r12c3.get()) + float(r14c3.get()) + float(r16c3.get()) + float(
                            r18c3.get()) + float(r20c3.get()) + float(r22c3.get()) + float(r24c3.get())
                        r25c3.insert(0, '{:.2f}'.format(float(res)))
                        quantity_paper_box.insert(0, '{:.0f}'.format(float(res)))
                        total_cost = float(res) * float(rate_box.get())
                        total_cost_box.insert(0, '{:.2f}'.format(float(total_cost)))

                    elif unit_box_cb.get() == 'kg':
                        try:
                            input_quantity = float(r1c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r1c1.delete(0, END)
                            r1c1.insert(0, 0)

                        try:
                            input_quantity = float(r3c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r3c1.delete(0, END)
                            r3c1.insert(0, 0)

                        try:
                            input_quantity = float(r5c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r5c1.delete(0, END)
                            r5c1.insert(0, 0)

                        try:
                            input_quantity = float(r7c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r7c1.delete(0, END)
                            r7c1.insert(0, 0)

                        try:
                            input_quantity = float(r9c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r9c1.delete(0, END)
                            r9c1.insert(0, 0)

                        try:
                            input_quantity = float(r11c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r11c1.delete(0, END)
                            r11c1.insert(0, 0)

                        try:
                            input_quantity = float(r13c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r13c1.delete(0, END)
                            r13c1.insert(0, 0)

                        try:
                            input_quantity = float(r15c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r15c1.delete(0, END)
                            r15c1.insert(0, 0)

                        try:
                            input_quantity = float(r17c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r17c1.delete(0, END)
                            r17c1.insert(0, 0)

                        try:
                            input_quantity = float(r19c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r19c1.delete(0, END)
                            r19c1.insert(0, 0)

                        try:
                            input_quantity = float(r21c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r21c1.delete(0, END)
                            r21c1.insert(0, 0)

                        try:
                            input_quantity = float(r23c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r23c1.delete(0, END)
                            r23c1.insert(0, 0)

                        try:
                            input_quantity = float(r1c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r1c2.delete(0, END)
                            r1c2.insert(0, 0)

                        try:
                            input_quantity = float(r3c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r3c2.delete(0, END)
                            r3c2.insert(0, 0)

                        try:
                            input_quantity = float(r5c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r5c2.delete(0, END)
                            r5c2.insert(0, 0)

                        try:
                            input_quantity = float(r7c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r7c2.delete(0, END)
                            r7c2.insert(0, 0)

                        try:
                            input_quantity = float(r9c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r9c2.delete(0, END)
                            r9c2.insert(0, 0)

                        try:
                            input_quantity = float(r11c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r11c2.delete(0, END)
                            r11c2.insert(0, 0)

                        try:
                            input_quantity = float(r13c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r13c2.delete(0, END)
                            r13c2.insert(0, 0)

                        try:
                            input_quantity = float(r15c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r15c2.delete(0, END)
                            r15c2.insert(0, 0)

                        try:
                            input_quantity = float(r17c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r17c2.delete(0, END)
                            r17c2.insert(0, 0)

                        try:
                            input_quantity = float(r19c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r19c2.delete(0, END)
                            r19c2.insert(0, 0)

                        try:
                            input_quantity = float(r21c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r21c2.delete(0, END)
                            r21c2.insert(0, 0)

                        try:
                            input_quantity = float(r23c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r23c2.delete(0, END)
                            r23c2.insert(0, 0)

                        try:
                            input_quantity = float(rate_box.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            rate_box.delete(0, END)
                            rate_box.insert(0, 0)

                        try:
                            input_quantity = float(quantity_kg.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            quantity_kg.delete(0, END)
                            quantity_kg.insert(0, 0)

                        c1r1 = float(r1c1.get())
                        c1r3 = float(r3c1.get())
                        c1r5 = float(r5c1.get())
                        c1r7 = float(r7c1.get())
                        c1r9 = float(r9c1.get())
                        c1r11 = float(r11c1.get())
                        c1r13 = float(r13c1.get())
                        c1r15 = float(r15c1.get())
                        c1r17 = float(r17c1.get())
                        c1r19 = float(r19c1.get())
                        c1r21 = float(r21c1.get())
                        c1r23 = float(r23c1.get())
                        c2r1 = float(r1c2.get())
                        c2r3 = float(r3c2.get())
                        c2r5 = float(r5c2.get())
                        c2r7 = float(r7c2.get())
                        c2r9 = float(r9c2.get())
                        c2r11 = float(r11c2.get())
                        c2r13 = float(r13c2.get())
                        c2r15 = float(r15c2.get())
                        c2r17 = float(r17c2.get())
                        c2r19 = float(r19c2.get())
                        c2r21 = float(r21c2.get())
                        c2r23 = float(r23c2.get())
                        new_kg = float(quantity_kg.get())

                        r2c3.insert(0, '{:.2f}'.format(c1r1 * c2r1))
                        r4c3.insert(0, '{:.2f}'.format(c1r3 * c2r3))
                        r6c3.insert(0, '{:.2f}'.format(c1r5 * c2r5))
                        r8c3.insert(0, '{:.2f}'.format(c1r7 * c2r7))
                        r10c3.insert(0, '{:.2f}'.format(c1r9 * c2r9))
                        r12c3.insert(0, '{:.2f}'.format(c1r11 * c2r11))
                        r14c3.insert(0, '{:.2f}'.format(c1r13 * c2r13))
                        r16c3.insert(0, '{:.2f}'.format(c1r15 * c2r15))
                        r18c3.insert(0, '{:.2f}'.format(c1r17 * c2r17))
                        r20c3.insert(0, '{:.2f}'.format(c1r19 * c2r19))
                        r22c3.insert(0, '{:.2f}'.format(c1r21 * c2r21))
                        r24c3.insert(0, '{:.2f}'.format(c1r23 * c2r23))
                        res = float(r2c3.get()) + float(r4c3.get()) + float(r6c3.get()) + float(r8c3.get()) + float(
                            r10c3.get()) + float(r12c3.get()) + float(r14c3.get()) + float(r16c3.get()) + float(
                            r18c3.get()) + float(r20c3.get()) + float(r22c3.get()) + float(r24c3.get())
                        r25c3.insert(0, '{:.2f}'.format(float(res)))
                        quantity_kg_o.insert(0, '{:.2f}'.format(float(res)))
                        quantity_paper_box.insert(0, '{:.0f}'.format(float(res) * new_kg))
                        new = res * new_kg
                        total_cost = float(new) * float(rate_box.get())
                        total_cost_box.insert(0, '{:.2f}'.format(float(total_cost)))

                    elif unit_box_cb.get() == 'm2':

                        try:
                            input_quantity = int(r1c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r1c1.delete(0, END)
                            r1c1.insert(0, 0)

                        try:
                            input_quantity = float(r4c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r4c1.delete(0, END)
                            r4c1.insert(0, 0)

                        try:
                            input_quantity = float(r7c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r7c1.delete(0, END)
                            r7c1.insert(0, 0)

                        try:
                            input_quantity = float(r10c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r10c1.delete(0, END)
                            r10c1.insert(0, 0)

                        try:
                            input_quantity = float(r13c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r13c1.delete(0, END)
                            r13c1.insert(0, 0)

                        try:
                            input_quantity = float(r16c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r16c1.delete(0, END)
                            r16c1.insert(0, 0)

                        try:
                            input_quantity = float(r19c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r19c1.delete(0, END)
                            r19c1.insert(0, 0)

                        try:
                            input_quantity = float(r22c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r22c1.delete(0, END)
                            r22c1.insert(0, 0)

                        try:
                            input_quantity = float(r1c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r1c2.delete(0, END)
                            r1c2.insert(0, 0)

                        try:
                            input_quantity = float(r2c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r2c2.delete(0, END)
                            r2c2.insert(0, 0)

                        try:
                            input_quantity = float(r4c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r4c2.delete(0, END)
                            r4c2.insert(0, 0)

                        try:
                            input_quantity = float(r5c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r5c2.delete(0, END)
                            r5c2.insert(0, 0)

                        try:
                            input_quantity = float(r7c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r7c2.delete(0, END)
                            r7c2.insert(0, 0)

                        try:
                            input_quantity = float(r8c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r8c2.delete(0, END)
                            r8c2.insert(0, 0)

                        try:
                            input_quantity = float(r10c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r10c2.delete(0, END)
                            r10c2.insert(0, 0)

                        try:
                            input_quantity = float(r11c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r11c2.delete(0, END)
                            r11c2.insert(0, 0)

                        try:
                            input_quantity = float(r13c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r13c2.delete(0, END)
                            r13c2.insert(0, 0)

                        try:
                            input_quantity = float(r14c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r14c2.delete(0, END)
                            r14c2.insert(0, 0)

                        try:
                            input_quantity = float(r16c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r16c2.delete(0, END)
                            r16c2.insert(0, 0)

                        try:
                            input_quantity = float(r17c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r17c2.delete(0, END)
                            r17c2.insert(0, 0)

                        try:
                            input_quantity = float(r19c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r19c2.delete(0, END)
                            r19c2.insert(0, 0)

                        try:
                            input_quantity = float(r20c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r20c2.delete(0, END)
                            r20c2.insert(0, 0)

                        try:
                            input_quantity = float(r22c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r22c2.delete(0, END)
                            r22c2.insert(0, 0)

                        try:
                            input_quantity = float(r23c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r23c2.delete(0, END)
                            r23c2.insert(0, 0)

                        try:
                            input_quantity = float(rate_box.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            rate_box.delete(0, END)
                            rate_box.insert(0, 0)

                        c1r1 = float(r1c1.get())
                        c1r4 = float(r4c1.get())
                        c1r7 = float(r7c1.get())
                        c1r10 = float(r10c1.get())
                        c1r13 = float(r13c1.get())
                        c1r16 = float(r16c1.get())
                        c1r19 = float(r19c1.get())
                        c1r22 = float(r22c1.get())

                        c2r1 = float(r1c2.get())
                        c2r2 = float(r2c2.get())
                        c2r4 = float(r4c2.get())
                        c2r5 = float(r5c2.get())
                        c2r7 = float(r7c2.get())
                        c2r8 = float(r8c2.get())
                        c2r10 = float(r10c2.get())
                        c2r11 = float(r11c2.get())
                        c2r13 = float(r13c2.get())
                        c2r14 = float(r14c2.get())
                        c2r16 = float(r16c2.get())
                        c2r17 = float(r17c2.get())
                        c2r19 = float(r19c2.get())
                        c2r20 = float(r20c2.get())
                        c2r22 = float(r22c2.get())
                        c2r23 = float(r23c2.get())

                        r3c3.insert(0, '{:.2f}'.format(c1r1 * (c2r1 * c2r2)))
                        r6c3.insert(0, '{:.2f}'.format(c1r4 * (c2r4 * c2r5)))
                        r9c3.insert(0, '{:.2f}'.format(c1r7 * (c2r7 * c2r8)))
                        r12c3.insert(0, '{:.2f}'.format(c1r10 * (c2r10 * c2r11)))
                        r15c3.insert(0, '{:.2f}'.format(c1r13 * (c2r13 * c2r14)))
                        r18c3.insert(0, '{:.2f}'.format(c1r16 * (c2r16 * c2r17)))
                        r21c3.insert(0, '{:.2f}'.format(c1r19 * (c2r19 * c2r20)))
                        r24c3.insert(0, '{:.2f}'.format(c1r22 * (c2r22 * c2r23)))
                        res = float(r3c3.get()) + float(r6c3.get()) + float(r9c3.get()) + float(r12c3.get()) + float(
                            r15c3.get()) + float(r18c3.get()) + float(r21c3.get()) + float(r24c3.get())
                        r25c3.insert(0, '{:.2f}'.format(float(res)))
                        quantity_paper_box.insert(0, '{:.0f}'.format(float(res)))
                        total_cost = float(res) * float(rate_box.get())
                        total_cost_box.insert(0, '{:.2f}'.format(float(total_cost)))

                    elif unit_box_cb.get() == 'm3':
                        try:
                            input_quantity = float(r1c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r1c1.delete(0, END)
                            r1c1.insert(0, 0)

                        try:
                            input_quantity = float(r5c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r5c1.delete(0, END)
                            r5c1.insert(0, 0)

                        try:
                            input_quantity = float(r9c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r9c1.delete(0, END)
                            r9c1.insert(0, 0)

                        try:
                            input_quantity = float(r13c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r13c1.delete(0, END)
                            r13c1.insert(0, 0)

                        try:
                            input_quantity = float(r17c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r17c1.delete(0, END)
                            r17c1.insert(0, 0)

                        try:
                            input_quantity = float(r21c1.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r21c1.delete(0, END)
                            r21c1.insert(0, 0)

                        try:
                            input_quantity = float(r1c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r1c2.delete(0, END)
                            r1c2.insert(0, 0)

                        try:
                            input_quantity = float(r2c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r2c2.delete(0, END)
                            r2c2.insert(0, 0)

                        try:
                            input_quantity = float(r3c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r3c2.delete(0, END)
                            r3c2.insert(0, 0)

                        try:
                            input_quantity = float(r5c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r5c2.delete(0, END)
                            r5c2.insert(0, 0)

                        try:
                            input_quantity = float(r6c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r6c2.delete(0, END)
                            r6c2.insert(0, 0)

                        try:
                            input_quantity = float(r7c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r7c2.delete(0, END)
                            r7c2.insert(0, 0)

                        try:
                            input_quantity = float(r9c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r9c2.delete(0, END)
                            r9c2.insert(0, 0)

                        try:
                            input_quantity = float(r10c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r10c2.delete(0, END)
                            r10c2.insert(0, 0)

                        try:
                            input_quantity = float(r11c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r11c2.delete(0, END)
                            r11c2.insert(0, 0)

                        try:
                            input_quantity = float(r13c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r13c2.delete(0, END)
                            r13c2.insert(0, 0)

                        try:
                            input_quantity = float(r14c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r14c2.delete(0, END)
                            r14c2.insert(0, 0)

                        try:
                            input_quantity = float(r15c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r15c2.delete(0, END)
                            r15c2.insert(0, 0)

                        try:
                            input_quantity = float(r17c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r17c2.delete(0, END)
                            r17c2.insert(0, 0)

                        try:
                            input_quantity = float(r18c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r18c2.delete(0, END)
                            r18c2.insert(0, 0)

                        try:
                            input_quantity = float(r19c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r19c2.delete(0, END)
                            r19c2.insert(0, 0)

                        try:
                            input_quantity = float(r21c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r21c2.delete(0, END)
                            r21c2.insert(0, 0)

                        try:
                            input_quantity = float(r22c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r22c2.delete(0, END)
                            r22c2.insert(0, 0)

                        try:
                            input_quantity = float(r23c2.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            r23c2.delete(0, END)
                            r23c2.insert(0, 0)

                        try:
                            input_quantity = float(rate_box.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            rate_box.delete(0, END)
                            rate_box.insert(0, 0)

                        c1r1 = float(r1c1.get())
                        c1r5 = float(r5c1.get())
                        c1r9 = float(r9c1.get())
                        c1r13 = float(r13c1.get())
                        c1r17 = float(r17c1.get())
                        c1r21 = float(r21c1.get())

                        c2r1 = float(r1c2.get())
                        c2r2 = float(r2c2.get())
                        c2r3 = float(r3c2.get())
                        c2r5 = float(r5c2.get())
                        c2r6 = float(r6c2.get())
                        c2r7 = float(r7c2.get())
                        c2r9 = float(r9c2.get())
                        c2r10 = float(r10c2.get())
                        c2r11 = float(r11c2.get())
                        c2r13 = float(r13c2.get())
                        c2r14 = float(r14c2.get())
                        c2r15 = float(r15c2.get())
                        c2r17 = float(r17c2.get())
                        c2r18 = float(r18c2.get())
                        c2r19 = float(r19c2.get())
                        c2r21 = float(r21c2.get())
                        c2r22 = float(r22c2.get())
                        c2r23 = float(r23c2.get())

                        r4c3.insert(0, '{:.2f}'.format(c1r1 * (c2r1 * c2r2 * c2r3)))
                        r8c3.insert(0, '{:.2f}'.format(c1r5 * (c2r5 * c2r6 * c2r7)))
                        r12c3.insert(0, '{:.2f}'.format(c1r9 * (c2r9 * c2r10 * c2r11)))
                        r16c3.insert(0, '{:.2f}'.format(c1r13 * (c2r13 * c2r14 * c2r15)))
                        r20c3.insert(0, '{:.2f}'.format(c1r17 * (c2r17 * c2r18 * c2r19)))
                        r24c3.insert(0, '{:.2f}'.format(c1r21 * (c2r21 * c2r22 * c2r23)))
                        res = float(r4c3.get()) + float(r8c3.get()) + float(r12c3.get()) + float(r16c3.get()) + float(
                            r20c3.get()) + float(r24c3.get())
                        r25c3.insert(0, '{:.2f}'.format(float(res)))
                        quantity_paper_box.insert(0, '{:.0f}'.format(float(res)))
                        total_cost = float(res) * float(rate_box.get())
                        total_cost_box.insert(0, '{:.2f}'.format(float(total_cost)))

                    elif unit_box_cb.get() == 'item':
                        try:
                            input_quantity = float(rate_box.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            rate_box.delete(0, END)
                            rate_box.insert(0, 0)

                        quantity_paper_box.insert(0, '{:.0f}'.format(0))
                        total_cost = float(rate_box.get())
                        total_cost_box.insert(0, '{:.2f}'.format(float(total_cost)))

                    else:
                        try:
                            input_quantity = float(rate_box.get())
                        except ValueError:
                            messagebox.showerror(message="NUMERIC INPUT ONLY")
                            rate_box.delete(0, END)
                        quantity_paper_box.insert(0, '{:.0f}'.format(0))
                        rate_box.insert(0, 0)
                        total_cost_box.insert(0, '{:.2f}'.format(0))

                def clear_calculat():
                    quantity_kg.delete(0, END)
                    quantity_kg_o.delete(0, END)
                    r1c1.delete(0, END)
                    r2c1.delete(0, END)
                    r3c1.delete(0, END)
                    r4c1.delete(0, END)
                    r5c1.delete(0, END)
                    r6c1.delete(0, END)
                    r7c1.delete(0, END)
                    r8c1.delete(0, END)
                    r9c1.delete(0, END)
                    r10c1.delete(0, END)
                    r11c1.delete(0, END)
                    r12c1.delete(0, END)
                    r13c1.delete(0, END)
                    r14c1.delete(0, END)
                    r15c1.delete(0, END)
                    r16c1.delete(0, END)
                    r17c1.delete(0, END)
                    r18c1.delete(0, END)
                    r19c1.delete(0, END)
                    r20c1.delete(0, END)
                    r21c1.delete(0, END)
                    r22c1.delete(0, END)
                    r23c1.delete(0, END)
                    r24c1.delete(0, END)

                    r1c2.delete(0, END)
                    r2c2.delete(0, END)
                    r3c2.delete(0, END)
                    r4c2.delete(0, END)
                    r5c2.delete(0, END)
                    r6c2.delete(0, END)
                    r7c2.delete(0, END)
                    r8c2.delete(0, END)
                    r9c2.delete(0, END)
                    r10c2.delete(0, END)
                    r11c2.delete(0, END)
                    r12c2.delete(0, END)
                    r13c2.delete(0, END)
                    r14c2.delete(0, END)
                    r15c2.delete(0, END)
                    r16c2.delete(0, END)
                    r17c2.delete(0, END)
                    r18c2.delete(0, END)
                    r19c2.delete(0, END)
                    r20c2.delete(0, END)
                    r21c2.delete(0, END)
                    r22c2.delete(0, END)
                    r23c2.delete(0, END)
                    r24c2.delete(0, END)

                    r1c3.delete(0, END)
                    r2c3.delete(0, END)
                    r3c3.delete(0, END)
                    r4c3.delete(0, END)
                    r5c3.delete(0, END)
                    r6c3.delete(0, END)
                    r7c3.delete(0, END)
                    r8c3.delete(0, END)
                    r9c3.delete(0, END)
                    r10c3.delete(0, END)
                    r11c3.delete(0, END)
                    r12c3.delete(0, END)
                    r13c3.delete(0, END)
                    r14c3.delete(0, END)
                    r15c3.delete(0, END)
                    r16c3.delete(0, END)
                    r17c3.delete(0, END)
                    r18c3.delete(0, END)
                    r19c3.delete(0, END)
                    r20c3.delete(0, END)
                    r21c3.delete(0, END)
                    r22c3.delete(0, END)
                    r23c3.delete(0, END)
                    r24c3.delete(0, END)
                    r25c3.delete(0, END)
                    rate_box.delete(0,END)

                    quantity_kg.insert(0, 0)
                    quantity_kg_o.insert(0, 0)
                    r1c1.insert(0, 0)
                    r2c1.insert(0, 0)
                    r3c1.insert(0, 0)
                    r4c1.insert(0, 0)
                    r5c1.insert(0, 0)
                    r6c1.insert(0, 0)
                    r7c1.insert(0, 0)
                    r8c1.insert(0, 0)
                    r9c1.insert(0, 0)
                    r10c1.insert(0, 0)
                    r11c1.insert(0, 0)
                    r12c1.insert(0, 0)
                    r13c1.insert(0, 0)
                    r14c1.insert(0, 0)
                    r15c1.insert(0, 0)
                    r16c1.insert(0, 0)
                    r17c1.insert(0, 0)
                    r18c1.insert(0, 0)
                    r19c1.insert(0, 0)
                    r20c1.insert(0, 0)
                    r21c1.insert(0, 0)
                    r22c1.insert(0, 0)
                    r23c1.insert(0, 0)

                    r1c2.insert(0, 0)
                    r2c2.insert(0, 0)
                    r3c2.insert(0, 0)
                    r4c2.insert(0, 0)
                    r5c2.insert(0, 0)
                    r6c2.insert(0, 0)
                    r7c2.insert(0, 0)
                    r8c2.insert(0, 0)
                    r9c2.insert(0, 0)
                    r10c2.insert(0, 0)
                    r11c2.insert(0, 0)
                    r12c2.insert(0, 0)
                    r13c2.insert(0, 0)
                    r14c2.insert(0, 0)
                    r15c2.insert(0, 0)
                    r16c2.insert(0, 0)
                    r17c2.insert(0, 0)
                    r18c2.insert(0, 0)
                    r19c2.insert(0, 0)
                    r20c2.insert(0, 0)
                    r21c2.insert(0, 0)
                    r22c2.insert(0, 0)
                    r23c2.insert(0, 0)
                    rate_box.insert(0, 0)

                    r1c3.insert(0, '')
                    r2c3.insert(0, '')
                    r3c3.insert(0, '')
                    r4c3.insert(0, '')
                    r5c3.insert(0, '')
                    r6c3.insert(0, '')
                    r7c3.insert(0, '')
                    r8c3.insert(0, '')
                    r9c3.insert(0, '')
                    r10c3.insert(0, '')
                    r11c3.insert(0, '')
                    r12c3.insert(0, '')
                    r13c3.insert(0, '')
                    r14c3.insert(0, '')
                    r15c3.insert(0, '')
                    r16c3.insert(0, '')
                    r17c3.insert(0, '')
                    r18c3.insert(0, '')
                    r19c3.insert(0, '')
                    r20c3.insert(0, '')
                    r21c3.insert(0, '')
                    r22c3.insert(0, '')
                    r23c3.insert(0, '')
                    r24c3.insert(0, '')
                    r25c3.insert(0, '')

                def unit_confirm():
                    calculation_frame.place(x=10, y=131)
                    covert_frame.place(x=304, y=595)
                    r1c1.delete(0, END)
                    r2c1.delete(0, END)
                    r3c1.delete(0, END)
                    r4c1.delete(0, END)
                    r5c1.delete(0, END)
                    r6c1.delete(0, END)
                    r7c1.delete(0, END)
                    r8c1.delete(0, END)
                    r9c1.delete(0, END)
                    r10c1.delete(0, END)
                    r11c1.delete(0, END)
                    r12c1.delete(0, END)
                    r13c1.delete(0, END)
                    r14c1.delete(0, END)
                    r15c1.delete(0, END)
                    r16c1.delete(0, END)
                    r17c1.delete(0, END)
                    r18c1.delete(0, END)
                    r19c1.delete(0, END)
                    r20c1.delete(0, END)
                    r21c1.delete(0, END)
                    r22c1.delete(0, END)
                    r23c1.delete(0, END)
                    r24c1.delete(0, END)

                    r1c2.delete(0, END)
                    r2c2.delete(0, END)
                    r3c2.delete(0, END)
                    r4c2.delete(0, END)
                    r5c2.delete(0, END)
                    r6c2.delete(0, END)
                    r7c2.delete(0, END)
                    r8c2.delete(0, END)
                    r9c2.delete(0, END)
                    r10c2.delete(0, END)
                    r11c2.delete(0, END)
                    r12c2.delete(0, END)
                    r13c2.delete(0, END)
                    r14c2.delete(0, END)
                    r15c2.delete(0, END)
                    r16c2.delete(0, END)
                    r17c2.delete(0, END)
                    r18c2.delete(0, END)
                    r19c2.delete(0, END)
                    r20c2.delete(0, END)
                    r21c2.delete(0, END)
                    r22c2.delete(0, END)
                    r23c2.delete(0, END)
                    r24c2.delete(0, END)

                    r1c3.delete(0, END)
                    r2c3.delete(0, END)
                    r3c3.delete(0, END)
                    r4c3.delete(0, END)
                    r5c3.delete(0, END)
                    r6c3.delete(0, END)
                    r7c3.delete(0, END)
                    r8c3.delete(0, END)
                    r9c3.delete(0, END)
                    r10c3.delete(0, END)
                    r11c3.delete(0, END)
                    r12c3.delete(0, END)
                    r13c3.delete(0, END)
                    r14c3.delete(0, END)
                    r15c3.delete(0, END)
                    r16c3.delete(0, END)
                    r17c3.delete(0, END)
                    r18c3.delete(0, END)
                    r19c3.delete(0, END)
                    r20c3.delete(0, END)
                    r21c3.delete(0, END)
                    r22c3.delete(0, END)
                    r23c3.delete(0, END)
                    r24c3.delete(0, END)
                    r25c3.delete(0, END)
                    rate_box.delete(0, END)
                    quantity_paper_box.delete(0, END)
                    total_cost_box.delete(0, END)
                    quantity_kg.delete(0, END)
                    quantity_kg_o.delete(0, END)

                    # for,m,kg and no
                    if unit_box_cb.get() == 'm' or unit_box_cb.get() == 'no':
                        item_label.place_forget()
                        r1c1.insert(0, 0)
                        r2c1.insert(0, 0)
                        r3c1.insert(0, 0)
                        r4c1.insert(0, 0)
                        r5c1.insert(0, 0)
                        r6c1.insert(0, 0)
                        r7c1.insert(0, 0)
                        r8c1.insert(0, 0)
                        r9c1.insert(0, 0)
                        r10c1.insert(0, 0)
                        r11c1.insert(0, 0)
                        r12c1.insert(0, 0)
                        r13c1.insert(0, 0)
                        r14c1.insert(0, 0)
                        r15c1.insert(0, 0)
                        r16c1.insert(0, 0)
                        r17c1.insert(0, 0)
                        r18c1.insert(0, 0)
                        r19c1.insert(0, 0)
                        r20c1.insert(0, 0)
                        r21c1.insert(0, 0)
                        r22c1.insert(0, 0)
                        r23c1.insert(0, 0)

                        r1c2.insert(0, 0)
                        r2c2.insert(0, 0)
                        r3c2.insert(0, 0)
                        r4c2.insert(0, 0)
                        r5c2.insert(0, 0)
                        r6c2.insert(0, 0)
                        r7c2.insert(0, 0)
                        r8c2.insert(0, 0)
                        r9c2.insert(0, 0)
                        r10c2.insert(0, 0)
                        r11c2.insert(0, 0)
                        r12c2.insert(0, 0)
                        r13c2.insert(0, 0)
                        r14c2.insert(0, 0)
                        r15c2.insert(0, 0)
                        r16c2.insert(0, 0)
                        r17c2.insert(0, 0)
                        r18c2.insert(0, 0)
                        r19c2.insert(0, 0)
                        r20c2.insert(0, 0)
                        r21c2.insert(0, 0)
                        r22c2.insert(0, 0)
                        r23c2.insert(0, 0)
                        rate_box.insert(0, 0)

                        r1c3.insert(0, '')
                        r2c3.insert(0, '')
                        r3c3.insert(0, '')
                        r4c3.insert(0, '')
                        r5c3.insert(0, '')
                        r6c3.insert(0, '')
                        r7c3.insert(0, '')
                        r8c3.insert(0, '')
                        r9c3.insert(0, '')
                        r10c3.insert(0, '')
                        r11c3.insert(0, '')
                        r12c3.insert(0, '')
                        r13c3.insert(0, '')
                        r14c3.insert(0, '')
                        r15c3.insert(0, '')
                        r16c3.insert(0, '')
                        r17c3.insert(0, '')
                        r18c3.insert(0, '')
                        r19c3.insert(0, '')
                        r20c3.insert(0, '')
                        r21c3.insert(0, '')
                        r22c3.insert(0, '')
                        r23c3.insert(0, '')
                        r24c3.insert(0, '')
                        r25c3.insert(0, '')

                        # reload entre box
                        r1c1.grid(row=0, column=0)
                        r2c1.grid(row=1, column=0)
                        r3c1.grid(row=2, column=0)
                        r4c1.grid(row=3, column=0)
                        r5c1.grid(row=4, column=0)
                        r6c1.grid(row=5, column=0)
                        r7c1.grid(row=6, column=0)
                        r8c1.grid(row=7, column=0)
                        r9c1.grid(row=8, column=0)
                        r10c1.grid(row=9, column=0)
                        r11c1.grid(row=10, column=0)
                        r12c1.grid(row=11, column=0)
                        r13c1.grid(row=12, column=0)
                        r14c1.grid(row=13, column=0)
                        r15c1.grid(row=14, column=0)
                        r16c1.grid(row=15, column=0)
                        r17c1.grid(row=16, column=0)
                        r18c1.grid(row=17, column=0)
                        r19c1.grid(row=18, column=0)
                        r20c1.grid(row=19, column=0)
                        r21c1.grid(row=20, column=0)
                        r22c1.grid(row=21, column=0)
                        r23c1.grid(row=22, column=0)
                        r24c1.grid(row=23, column=0)

                        r1c2.grid(row=0, column=1)
                        r2c2.grid(row=1, column=1)
                        r3c2.grid(row=2, column=1)
                        r4c2.grid(row=3, column=1)
                        r5c2.grid(row=4, column=1)
                        r6c2.grid(row=5, column=1)
                        r7c2.grid(row=6, column=1)
                        r8c2.grid(row=7, column=1)
                        r9c2.grid(row=8, column=1)
                        r10c2.grid(row=9, column=1)
                        r11c2.grid(row=10, column=1)
                        r12c2.grid(row=11, column=1)
                        r13c2.grid(row=12, column=1)
                        r14c2.grid(row=13, column=1)
                        r15c2.grid(row=14, column=1)
                        r16c2.grid(row=15, column=1)
                        r17c2.grid(row=16, column=1)
                        r18c2.grid(row=17, column=1)
                        r19c2.grid(row=18, column=1)
                        r20c2.grid(row=19, column=1)
                        r21c2.grid(row=20, column=1)
                        r22c2.grid(row=21, column=1)
                        r23c2.grid(row=22, column=1)
                        r24c2.grid(row=23, column=1)

                        r1c3.grid(row=0, column=2)
                        r2c3.grid(row=1, column=2)
                        r3c3.grid(row=2, column=2)
                        r4c3.grid(row=3, column=2)
                        r5c3.grid(row=4, column=2)
                        r6c3.grid(row=5, column=2)
                        r7c3.grid(row=6, column=2)
                        r8c3.grid(row=7, column=2)
                        r9c3.grid(row=8, column=2)
                        r10c3.grid(row=9, column=2)
                        r11c3.grid(row=10, column=2)
                        r12c3.grid(row=11, column=2)
                        r13c3.grid(row=12, column=2)
                        r14c3.grid(row=13, column=2)
                        r15c3.grid(row=14, column=2)
                        r16c3.grid(row=15, column=2)
                        r17c3.grid(row=16, column=2)
                        r18c3.grid(row=17, column=2)
                        r19c3.grid(row=18, column=2)
                        r20c3.grid(row=19, column=2)
                        r21c3.grid(row=20, column=2)
                        r22c3.grid(row=21, column=2)
                        r23c3.grid(row=22, column=2)
                        r24c3.grid(row=23, column=2)
                        r25c3.grid(row=24, column=2)
                        # remove entre box

                        r1c3.grid_remove()
                        r3c3.grid_remove()
                        r5c3.grid_remove()
                        r7c3.grid_remove()
                        r9c3.grid_remove()
                        r11c3.grid_remove()
                        r13c3.grid_remove()
                        r15c3.grid_remove()
                        r17c3.grid_remove()
                        r19c3.grid_remove()
                        r21c3.grid_remove()
                        r23c3.grid_remove()

                        r2c1.grid_remove()
                        r4c1.grid_remove()
                        r6c1.grid_remove()
                        r8c1.grid_remove()
                        r10c1.grid_remove()
                        r12c1.grid_remove()
                        r14c1.grid_remove()
                        r16c1.grid_remove()
                        r18c1.grid_remove()
                        r20c1.grid_remove()
                        r22c1.grid_remove()
                        r24c1.grid_remove()

                        r2c2.grid_remove()
                        r4c2.grid_remove()
                        r6c2.grid_remove()
                        r8c2.grid_remove()
                        r10c2.grid_remove()
                        r12c2.grid_remove()
                        r14c2.grid_remove()
                        r16c2.grid_remove()
                        r18c2.grid_remove()
                        r20c2.grid_remove()
                        r22c2.grid_remove()
                        r24c2.grid_remove()
                        covert_frame.place_forget()

                    elif unit_box_cb.get() == 'kg':
                        item_label.place_forget()
                        covert_frame.place(x=304, y=595)
                        quantity_kg.insert(0, 0)
                        quantity_kg_o.insert(0, 0)
                        r1c1.insert(0, 0)
                        r2c1.insert(0, 0)
                        r3c1.insert(0, 0)
                        r4c1.insert(0, 0)
                        r5c1.insert(0, 0)
                        r6c1.insert(0, 0)
                        r7c1.insert(0, 0)
                        r8c1.insert(0, 0)
                        r9c1.insert(0, 0)
                        r10c1.insert(0, 0)
                        r11c1.insert(0, 0)
                        r12c1.insert(0, 0)
                        r13c1.insert(0, 0)
                        r14c1.insert(0, 0)
                        r15c1.insert(0, 0)
                        r16c1.insert(0, 0)
                        r17c1.insert(0, 0)
                        r18c1.insert(0, 0)
                        r19c1.insert(0, 0)
                        r20c1.insert(0, 0)
                        r21c1.insert(0, 0)
                        r22c1.insert(0, 0)
                        r23c1.insert(0, 0)

                        r1c2.insert(0, 0)
                        r2c2.insert(0, 0)
                        r3c2.insert(0, 0)
                        r4c2.insert(0, 0)
                        r5c2.insert(0, 0)
                        r6c2.insert(0, 0)
                        r7c2.insert(0, 0)
                        r8c2.insert(0, 0)
                        r9c2.insert(0, 0)
                        r10c2.insert(0, 0)
                        r11c2.insert(0, 0)
                        r12c2.insert(0, 0)
                        r13c2.insert(0, 0)
                        r14c2.insert(0, 0)
                        r15c2.insert(0, 0)
                        r16c2.insert(0, 0)
                        r17c2.insert(0, 0)
                        r18c2.insert(0, 0)
                        r19c2.insert(0, 0)
                        r20c2.insert(0, 0)
                        r21c2.insert(0, 0)
                        r22c2.insert(0, 0)
                        r23c2.insert(0, 0)
                        rate_box.insert(0, 0)

                        r1c3.insert(0, '')
                        r2c3.insert(0, '')
                        r3c3.insert(0, '')
                        r4c3.insert(0, '')
                        r5c3.insert(0, '')
                        r6c3.insert(0, '')
                        r7c3.insert(0, '')
                        r8c3.insert(0, '')
                        r9c3.insert(0, '')
                        r10c3.insert(0, '')
                        r11c3.insert(0, '')
                        r12c3.insert(0, '')
                        r13c3.insert(0, '')
                        r14c3.insert(0, '')
                        r15c3.insert(0, '')
                        r16c3.insert(0, '')
                        r17c3.insert(0, '')
                        r18c3.insert(0, '')
                        r19c3.insert(0, '')
                        r20c3.insert(0, '')
                        r21c3.insert(0, '')
                        r22c3.insert(0, '')
                        r23c3.insert(0, '')
                        r24c3.insert(0, '')
                        r25c3.insert(0, '')
                        # reload entre box
                        r1c1.grid(row=0, column=0)
                        r2c1.grid(row=1, column=0)
                        r3c1.grid(row=2, column=0)
                        r4c1.grid(row=3, column=0)
                        r5c1.grid(row=4, column=0)
                        r6c1.grid(row=5, column=0)
                        r7c1.grid(row=6, column=0)
                        r8c1.grid(row=7, column=0)
                        r9c1.grid(row=8, column=0)
                        r10c1.grid(row=9, column=0)
                        r11c1.grid(row=10, column=0)
                        r12c1.grid(row=11, column=0)
                        r13c1.grid(row=12, column=0)
                        r14c1.grid(row=13, column=0)
                        r15c1.grid(row=14, column=0)
                        r16c1.grid(row=15, column=0)
                        r17c1.grid(row=16, column=0)
                        r18c1.grid(row=17, column=0)
                        r19c1.grid(row=18, column=0)
                        r20c1.grid(row=19, column=0)
                        r21c1.grid(row=20, column=0)
                        r22c1.grid(row=21, column=0)
                        r23c1.grid(row=22, column=0)
                        r24c1.grid(row=23, column=0)

                        r1c2.grid(row=0, column=1)
                        r2c2.grid(row=1, column=1)
                        r3c2.grid(row=2, column=1)
                        r4c2.grid(row=3, column=1)
                        r5c2.grid(row=4, column=1)
                        r6c2.grid(row=5, column=1)
                        r7c2.grid(row=6, column=1)
                        r8c2.grid(row=7, column=1)
                        r9c2.grid(row=8, column=1)
                        r10c2.grid(row=9, column=1)
                        r11c2.grid(row=10, column=1)
                        r12c2.grid(row=11, column=1)
                        r13c2.grid(row=12, column=1)
                        r14c2.grid(row=13, column=1)
                        r15c2.grid(row=14, column=1)
                        r16c2.grid(row=15, column=1)
                        r17c2.grid(row=16, column=1)
                        r18c2.grid(row=17, column=1)
                        r19c2.grid(row=18, column=1)
                        r20c2.grid(row=19, column=1)
                        r21c2.grid(row=20, column=1)
                        r22c2.grid(row=21, column=1)
                        r23c2.grid(row=22, column=1)
                        r24c2.grid(row=23, column=1)

                        r1c3.grid(row=0, column=2)
                        r2c3.grid(row=1, column=2)
                        r3c3.grid(row=2, column=2)
                        r4c3.grid(row=3, column=2)
                        r5c3.grid(row=4, column=2)
                        r6c3.grid(row=5, column=2)
                        r7c3.grid(row=6, column=2)
                        r8c3.grid(row=7, column=2)
                        r9c3.grid(row=8, column=2)
                        r10c3.grid(row=9, column=2)
                        r11c3.grid(row=10, column=2)
                        r12c3.grid(row=11, column=2)
                        r13c3.grid(row=12, column=2)
                        r14c3.grid(row=13, column=2)
                        r15c3.grid(row=14, column=2)
                        r16c3.grid(row=15, column=2)
                        r17c3.grid(row=16, column=2)
                        r18c3.grid(row=17, column=2)
                        r19c3.grid(row=18, column=2)
                        r20c3.grid(row=19, column=2)
                        r21c3.grid(row=20, column=2)
                        r22c3.grid(row=21, column=2)
                        r23c3.grid(row=22, column=2)
                        r24c3.grid(row=23, column=2)
                        r25c3.grid(row=24, column=2)
                        # remove entre box

                        r1c3.grid_remove()
                        r3c3.grid_remove()
                        r5c3.grid_remove()
                        r7c3.grid_remove()
                        r9c3.grid_remove()
                        r11c3.grid_remove()
                        r13c3.grid_remove()
                        r15c3.grid_remove()
                        r17c3.grid_remove()
                        r19c3.grid_remove()
                        r21c3.grid_remove()
                        r23c3.grid_remove()

                        r2c1.grid_remove()
                        r4c1.grid_remove()
                        r6c1.grid_remove()
                        r8c1.grid_remove()
                        r10c1.grid_remove()
                        r12c1.grid_remove()
                        r14c1.grid_remove()
                        r16c1.grid_remove()
                        r18c1.grid_remove()
                        r20c1.grid_remove()
                        r22c1.grid_remove()
                        r24c1.grid_remove()

                        r2c2.grid_remove()
                        r4c2.grid_remove()
                        r6c2.grid_remove()
                        r8c2.grid_remove()
                        r10c2.grid_remove()
                        r12c2.grid_remove()
                        r14c2.grid_remove()
                        r16c2.grid_remove()
                        r18c2.grid_remove()
                        r20c2.grid_remove()
                        r22c2.grid_remove()
                        r24c2.grid_remove()


                    elif unit_box_cb.get() == 'm2':
                        covert_frame.place_forget()
                        r1c1.insert(0, 0)
                        r2c1.insert(0, 0)
                        r3c1.insert(0, 0)
                        r4c1.insert(0, 0)
                        r5c1.insert(0, 0)
                        r6c1.insert(0, 0)
                        r7c1.insert(0, 0)
                        r8c1.insert(0, 0)
                        r9c1.insert(0, 0)
                        r10c1.insert(0, 0)
                        r11c1.insert(0, 0)
                        r12c1.insert(0, 0)
                        r13c1.insert(0, 0)
                        r14c1.insert(0, 0)
                        r15c1.insert(0, 0)
                        r16c1.insert(0, 0)
                        r17c1.insert(0, 0)
                        r18c1.insert(0, 0)
                        r19c1.insert(0, 0)
                        r20c1.insert(0, 0)
                        r21c1.insert(0, 0)
                        r22c1.insert(0, 0)
                        r23c1.insert(0, 0)

                        r1c2.insert(0, 0)
                        r2c2.insert(0, 0)
                        r3c2.insert(0, 0)
                        r4c2.insert(0, 0)
                        r5c2.insert(0, 0)
                        r6c2.insert(0, 0)
                        r7c2.insert(0, 0)
                        r8c2.insert(0, 0)
                        r9c2.insert(0, 0)
                        r10c2.insert(0, 0)
                        r11c2.insert(0, 0)
                        r12c2.insert(0, 0)
                        r13c2.insert(0, 0)
                        r14c2.insert(0, 0)
                        r15c2.insert(0, 0)
                        r16c2.insert(0, 0)
                        r17c2.insert(0, 0)
                        r18c2.insert(0, 0)
                        r19c2.insert(0, 0)
                        r20c2.insert(0, 0)
                        r21c2.insert(0, 0)
                        r22c2.insert(0, 0)
                        r23c2.insert(0, 0)

                        r1c3.insert(0, '')
                        r2c3.insert(0, '')
                        r3c3.insert(0, '')
                        r4c3.insert(0, '')
                        r5c3.insert(0, '')
                        r6c3.insert(0, '')
                        r7c3.insert(0, '')
                        r8c3.insert(0, '')
                        r9c3.insert(0, '')
                        r10c3.insert(0, '')
                        r11c3.insert(0, '')
                        r12c3.insert(0, '')
                        r13c3.insert(0, '')
                        r14c3.insert(0, '')
                        r15c3.insert(0, '')
                        r16c3.insert(0, '')
                        r17c3.insert(0, '')
                        r18c3.insert(0, '')
                        r19c3.insert(0, '')
                        r20c3.insert(0, '')
                        r21c3.insert(0, '')
                        r22c3.insert(0, '')
                        r23c3.insert(0, '')
                        r24c3.insert(0, '')
                        r25c3.insert(0, '')
                        rate_box.insert(0, 0)
                        # reload entry box
                        r1c1.grid(row=0, column=0)
                        r2c1.grid(row=1, column=0)
                        r3c1.grid(row=2, column=0)
                        r4c1.grid(row=3, column=0)
                        r5c1.grid(row=4, column=0)
                        r6c1.grid(row=5, column=0)
                        r7c1.grid(row=6, column=0)
                        r8c1.grid(row=7, column=0)
                        r9c1.grid(row=8, column=0)
                        r10c1.grid(row=9, column=0)
                        r11c1.grid(row=10, column=0)
                        r12c1.grid(row=11, column=0)
                        r13c1.grid(row=12, column=0)
                        r14c1.grid(row=13, column=0)
                        r15c1.grid(row=14, column=0)
                        r16c1.grid(row=15, column=0)
                        r17c1.grid(row=16, column=0)
                        r18c1.grid(row=17, column=0)
                        r19c1.grid(row=18, column=0)
                        r20c1.grid(row=19, column=0)
                        r21c1.grid(row=20, column=0)
                        r22c1.grid(row=21, column=0)
                        r23c1.grid(row=22, column=0)
                        r24c1.grid(row=23, column=0)

                        r1c2.grid(row=0, column=1)
                        r2c2.grid(row=1, column=1)
                        r3c2.grid(row=2, column=1)
                        r4c2.grid(row=3, column=1)
                        r5c2.grid(row=4, column=1)
                        r6c2.grid(row=5, column=1)
                        r7c2.grid(row=6, column=1)
                        r8c2.grid(row=7, column=1)
                        r9c2.grid(row=8, column=1)
                        r10c2.grid(row=9, column=1)
                        r11c2.grid(row=10, column=1)
                        r12c2.grid(row=11, column=1)
                        r13c2.grid(row=12, column=1)
                        r14c2.grid(row=13, column=1)
                        r15c2.grid(row=14, column=1)
                        r16c2.grid(row=15, column=1)
                        r17c2.grid(row=16, column=1)
                        r18c2.grid(row=17, column=1)
                        r19c2.grid(row=18, column=1)
                        r20c2.grid(row=19, column=1)
                        r21c2.grid(row=20, column=1)
                        r22c2.grid(row=21, column=1)
                        r23c2.grid(row=22, column=1)
                        r24c2.grid(row=23, column=1)

                        r1c3.grid(row=0, column=2)
                        r2c3.grid(row=1, column=2)
                        r3c3.grid(row=2, column=2)
                        r4c3.grid(row=3, column=2)
                        r5c3.grid(row=4, column=2)
                        r6c3.grid(row=5, column=2)
                        r7c3.grid(row=6, column=2)
                        r8c3.grid(row=7, column=2)
                        r9c3.grid(row=8, column=2)
                        r10c3.grid(row=9, column=2)
                        r11c3.grid(row=10, column=2)
                        r12c3.grid(row=11, column=2)
                        r13c3.grid(row=12, column=2)
                        r14c3.grid(row=13, column=2)
                        r15c3.grid(row=14, column=2)
                        r16c3.grid(row=15, column=2)
                        r17c3.grid(row=16, column=2)
                        r18c3.grid(row=17, column=2)
                        r19c3.grid(row=18, column=2)
                        r20c3.grid(row=19, column=2)
                        r21c3.grid(row=20, column=2)
                        r22c3.grid(row=21, column=2)
                        r23c3.grid(row=22, column=2)
                        r24c3.grid(row=23, column=2)
                        r25c3.grid(row=24, column=2)
                        # remove entry box
                        r2c1.grid_remove()
                        r3c1.grid_remove()
                        r5c1.grid_remove()
                        r6c1.grid_remove()
                        r8c1.grid_remove()
                        r9c1.grid_remove()
                        r11c1.grid_remove()
                        r12c1.grid_remove()
                        r14c1.grid_remove()
                        r15c1.grid_remove()
                        r17c1.grid_remove()
                        r18c1.grid_remove()
                        r20c1.grid_remove()
                        r21c1.grid_remove()
                        r23c1.grid_remove()
                        r24c1.grid_remove()

                        r1c3.grid_remove()
                        r2c3.grid_remove()
                        r4c3.grid_remove()
                        r5c3.grid_remove()
                        r7c3.grid_remove()
                        r8c3.grid_remove()
                        r10c3.grid_remove()
                        r11c3.grid_remove()
                        r13c3.grid_remove()
                        r14c3.grid_remove()
                        r16c3.grid_remove()
                        r17c3.grid_remove()
                        r19c3.grid_remove()
                        r20c3.grid_remove()
                        r22c3.grid_remove()
                        r23c3.grid_remove()

                        r3c2.grid_remove()
                        r6c2.grid_remove()
                        r9c2.grid_remove()
                        r12c2.grid_remove()
                        r15c2.grid_remove()
                        r18c2.grid_remove()
                        r21c2.grid_remove()
                        r24c2.grid_remove()
                        item_label.place_forget()


                    elif unit_box_cb.get() == "m3":
                        covert_frame.place_forget()
                        rate_box.insert(0, 0)
                        r1c1.insert(0, 0)
                        r2c1.insert(0, 0)
                        r3c1.insert(0, 0)
                        r4c1.insert(0, 0)
                        r5c1.insert(0, 0)
                        r6c1.insert(0, 0)
                        r7c1.insert(0, 0)
                        r8c1.insert(0, 0)
                        r9c1.insert(0, 0)
                        r10c1.insert(0, 0)
                        r11c1.insert(0, 0)
                        r12c1.insert(0, 0)
                        r13c1.insert(0, 0)
                        r14c1.insert(0, 0)
                        r15c1.insert(0, 0)
                        r16c1.insert(0, 0)
                        r17c1.insert(0, 0)
                        r18c1.insert(0, 0)
                        r19c1.insert(0, 0)
                        r20c1.insert(0, 0)
                        r21c1.insert(0, 0)
                        r22c1.insert(0, 0)
                        r23c1.insert(0, 0)

                        r1c2.insert(0, 0)
                        r2c2.insert(0, 0)
                        r3c2.insert(0, 0)
                        r4c2.insert(0, 0)
                        r5c2.insert(0, 0)
                        r6c2.insert(0, 0)
                        r7c2.insert(0, 0)
                        r8c2.insert(0, 0)
                        r9c2.insert(0, 0)
                        r10c2.insert(0, 0)
                        r11c2.insert(0, 0)
                        r12c2.insert(0, 0)
                        r13c2.insert(0, 0)
                        r14c2.insert(0, 0)
                        r15c2.insert(0, 0)
                        r16c2.insert(0, 0)
                        r17c2.insert(0, 0)
                        r18c2.insert(0, 0)
                        r19c2.insert(0, 0)
                        r20c2.insert(0, 0)
                        r21c2.insert(0, 0)
                        r22c2.insert(0, 0)
                        r23c2.insert(0, 0)

                        r1c3.insert(0, '')
                        r2c3.insert(0, '')
                        r3c3.insert(0, '')
                        r4c3.insert(0, '')
                        r5c3.insert(0, '')
                        r6c3.insert(0, '')
                        r7c3.insert(0, '')
                        r8c3.insert(0, '')
                        r9c3.insert(0, '')
                        r10c3.insert(0, '')
                        r11c3.insert(0, '')
                        r12c3.insert(0, '')
                        r13c3.insert(0, '')
                        r14c3.insert(0, '')
                        r15c3.insert(0, '')
                        r16c3.insert(0, '')
                        r17c3.insert(0, '')
                        r18c3.insert(0, '')
                        r19c3.insert(0, '')
                        r20c3.insert(0, '')
                        r21c3.insert(0, '')
                        r22c3.insert(0, '')
                        r23c3.insert(0, '')
                        r24c3.insert(0, '')
                        r25c3.insert(0, '')
                        # reload entre box
                        r1c1.grid(row=0, column=0)
                        r2c1.grid(row=1, column=0)
                        r3c1.grid(row=2, column=0)
                        r4c1.grid(row=3, column=0)
                        r5c1.grid(row=4, column=0)
                        r6c1.grid(row=5, column=0)
                        r7c1.grid(row=6, column=0)
                        r8c1.grid(row=7, column=0)
                        r9c1.grid(row=8, column=0)
                        r10c1.grid(row=9, column=0)
                        r11c1.grid(row=10, column=0)
                        r12c1.grid(row=11, column=0)
                        r13c1.grid(row=12, column=0)
                        r14c1.grid(row=13, column=0)
                        r15c1.grid(row=14, column=0)
                        r16c1.grid(row=15, column=0)
                        r17c1.grid(row=16, column=0)
                        r18c1.grid(row=17, column=0)
                        r19c1.grid(row=18, column=0)
                        r20c1.grid(row=19, column=0)
                        r21c1.grid(row=20, column=0)
                        r22c1.grid(row=21, column=0)
                        r23c1.grid(row=22, column=0)
                        r24c1.grid(row=23, column=0)

                        r1c2.grid(row=0, column=1)
                        r2c2.grid(row=1, column=1)
                        r3c2.grid(row=2, column=1)
                        r4c2.grid(row=3, column=1)
                        r5c2.grid(row=4, column=1)
                        r6c2.grid(row=5, column=1)
                        r7c2.grid(row=6, column=1)
                        r8c2.grid(row=7, column=1)
                        r9c2.grid(row=8, column=1)
                        r10c2.grid(row=9, column=1)
                        r11c2.grid(row=10, column=1)
                        r12c2.grid(row=11, column=1)
                        r13c2.grid(row=12, column=1)
                        r14c2.grid(row=13, column=1)
                        r15c2.grid(row=14, column=1)
                        r16c2.grid(row=15, column=1)
                        r17c2.grid(row=16, column=1)
                        r18c2.grid(row=17, column=1)
                        r19c2.grid(row=18, column=1)
                        r20c2.grid(row=19, column=1)
                        r21c2.grid(row=20, column=1)
                        r22c2.grid(row=21, column=1)
                        r23c2.grid(row=22, column=1)
                        r24c2.grid(row=23, column=1)

                        r1c3.grid(row=0, column=2)
                        r2c3.grid(row=1, column=2)
                        r3c3.grid(row=2, column=2)
                        r4c3.grid(row=3, column=2)
                        r5c3.grid(row=4, column=2)
                        r6c3.grid(row=5, column=2)
                        r7c3.grid(row=6, column=2)
                        r8c3.grid(row=7, column=2)
                        r9c3.grid(row=8, column=2)
                        r10c3.grid(row=9, column=2)
                        r11c3.grid(row=10, column=2)
                        r12c3.grid(row=11, column=2)
                        r13c3.grid(row=12, column=2)
                        r14c3.grid(row=13, column=2)
                        r15c3.grid(row=14, column=2)
                        r16c3.grid(row=15, column=2)
                        r17c3.grid(row=16, column=2)
                        r18c3.grid(row=17, column=2)
                        r19c3.grid(row=18, column=2)
                        r20c3.grid(row=19, column=2)
                        r21c3.grid(row=20, column=2)
                        r22c3.grid(row=21, column=2)
                        r23c3.grid(row=22, column=2)
                        r24c3.grid(row=23, column=2)
                        r25c3.grid(row=24, column=2)

                        # remove entre box
                        covert_frame.place_forget()
                        r4c2.grid_remove()
                        r8c2.grid_remove()
                        r12c2.grid_remove()
                        r16c2.grid_remove()
                        r20c2.grid_remove()
                        r24c2.grid_remove()

                        r2c1.grid_remove()
                        r3c1.grid_remove()
                        r4c1.grid_remove()
                        r6c1.grid_remove()
                        r7c1.grid_remove()
                        r8c1.grid_remove()
                        r10c1.grid_remove()
                        r11c1.grid_remove()
                        r12c1.grid_remove()
                        r14c1.grid_remove()
                        r15c1.grid_remove()
                        r16c1.grid_remove()
                        r18c1.grid_remove()
                        r19c1.grid_remove()
                        r20c1.grid_remove()
                        r22c1.grid_remove()
                        r23c1.grid_remove()
                        r24c1.grid_remove()

                        r1c3.grid_remove()
                        r2c3.grid_remove()
                        r3c3.grid_remove()
                        r5c3.grid_remove()
                        r6c3.grid_remove()
                        r7c3.grid_remove()
                        r9c3.grid_remove()
                        r10c3.grid_remove()
                        r11c3.grid_remove()
                        r13c3.grid_remove()
                        r14c3.grid_remove()
                        r15c3.grid_remove()
                        r17c3.grid_remove()
                        r18c3.grid_remove()
                        r19c3.grid_remove()
                        r21c3.grid_remove()
                        r22c3.grid_remove()
                        r23c3.grid_remove()
                        item_label.place_forget()


                    elif unit_box_cb.get() == "item":
                        covert_frame.place_forget()
                        rate_box.insert(0, 0)
                        calculation_frame.place_forget()
                        item_label.place(x=130, y=300)
                        covert_frame.place_forget()

                    else:
                        covert_frame.place_forget()
                        calculation_frame.place_forget()
                        rate_box.insert(0, 0)
                        quantity_paper_box.insert(0, 0)
                        total_cost_box.insert(0, 0)

                def delete():
                    MsgBox = messagebox.askquestion(
                        message='PLEASE MAKE SURE YOU HAVE ALREADY SAVED THE COMPLECTED PROJECT UNDER ANOTHER FILE NAME BEFORE CLEARING THE ENTIRE EXCEL FILE. DO YOU WANT TO CLEAR THE ENTIRE EXCEL SHEET?')
                    if MsgBox == 'yes':
                        os.remove("TAKING_OFF.xlsx")

                def save():
                    filepath = "my_image.png"
                    ss = ImageGrab.grab(bbox=(785, 250, 1100, 817))
                    ss.save(filepath, "PNG")

                    MsgBox = messagebox.askquestion(message='DO YOU WANT TO SAVE THE DATA TO EXCEL ?')

                    if MsgBox == 'yes':
                        heading = heading_box.get()
                        description = description_box.get()
                        quantity = quantity_paper_box.get()
                        unit = unit_box_cb.get()
                        cost = rate_box.get()
                        total_cost = total_cost_box.get()
                        bill = bill_box.get()
                        project_name = job_box.get()
                        element = element_box.get()
                        slip_no = slip_box.get()
                        taker = taker_off_box.get()
                        sqd = sqd_box.get()
                        ch_sqd = ch_sqd_box.get()
                        red = red_box.get()
                        c1r1 = (r1c1.get())
                        c1r2 = (r2c1.get())
                        c1r3 = (r3c1.get())
                        c1r4 = (r4c1.get())
                        c1r5 = (r5c1.get())
                        c1r6 = (r6c1.get())
                        c1r7 = (r7c1.get())
                        c1r8 = (r8c1.get())
                        c1r9 = (r9c1.get())
                        c1r10 = (r10c1.get())
                        c1r11 = (r11c1.get())
                        c1r12 = (r12c1.get())
                        c1r13 = (r13c1.get())
                        c1r14 = (r14c1.get())
                        c1r15 = (r15c1.get())
                        c1r16 = (r16c1.get())
                        c1r17 = (r17c1.get())
                        c1r18 = (r18c1.get())
                        c1r19 = (r19c1.get())
                        c1r20 = (r20c1.get())
                        c1r21 = (r21c1.get())
                        c1r22 = (r22c1.get())
                        c1r23 = (r23c1.get())
                        c1r24 = (r24c1.get())
                        c2r1 = (r1c2.get())
                        c2r2 = (r2c2.get())
                        c2r3 = (r3c2.get())
                        c2r4 = (r4c2.get())
                        c2r5 = (r5c2.get())
                        c2r6 = (r6c2.get())
                        c2r7 = (r7c2.get())
                        c2r8 = (r8c2.get())
                        c2r9 = (r9c2.get())
                        c2r10 = (r10c2.get())
                        c2r11 = (r11c2.get())
                        c2r12 = (r12c2.get())
                        c2r13 = (r13c2.get())
                        c2r14 = (r14c2.get())
                        c2r15 = (r15c2.get())
                        c2r16 = (r16c2.get())
                        c2r17 = (r17c2.get())
                        c2r18 = (r18c2.get())
                        c2r19 = (r19c2.get())
                        c2r20 = (r20c2.get())
                        c2r21 = (r21c2.get())
                        c2r22 = (r22c2.get())
                        c2r23 = (r23c2.get())
                        c2r24 = (r24c2.get())
                        kg = (quantity_kg_o.get())
                        kg_new = (quantity_kg.get())
                        c3r1 = (r1c3.get())
                        c3r2 = (r2c3.get())
                        c3r3 = (r3c3.get())
                        c3r4 = (r4c3.get())
                        c3r5 = (r5c3.get())
                        c3r6 = (r6c3.get())
                        c3r7 = (r7c3.get())
                        c3r8 = (r8c3.get())
                        c3r9 = (r9c3.get())
                        c3r10 = (r10c3.get())
                        c3r11 = (r11c3.get())
                        c3r12 = (r12c3.get())
                        c3r13 = (r13c3.get())
                        c3r14 = (r14c3.get())
                        c3r15 = (r15c3.get())
                        c3r16 = (r16c3.get())
                        c3r17 = (r17c3.get())
                        c3r18 = (r18c3.get())
                        c3r19 = (r19c3.get())
                        c3r20 = (r20c3.get())
                        c3r21 = (r21c3.get())
                        c3r22 = (r22c3.get())
                        c3r23 = (r23c3.get())
                        c3r24 = (r24c3.get())
                        c3r25 = (r25c3.get())


                        if os.path.exists("TAKING_OFF.xlsx"):
                            file = load_workbook("TAKING_OFF.xlsx")
                            last_sheet = file.worksheets[-1]
                            index = int(last_sheet.title)
                            sheet = file.create_sheet('index')
                            sheet.title = str(index + 1)

                        else:
                            file = Workbook()
                            sheet = file.worksheets[0]
                            sheet.title = '1'

                        sheet.merge_cells('A1:B1')  # JOB Heading
                        sheet.merge_cells('C1:D1')  # BIL NO Heading
                        sheet.merge_cells('E1:F1')  # ELEMENT NO HEADING
                        sheet.merge_cells('G1:H1')  # SLIP NO HEADING
                        sheet.merge_cells('A2:B2')  # job data
                        sheet.merge_cells('C2:D2')  # bill no data
                        sheet.merge_cells('E2:F2')  # element no data
                        sheet.merge_cells('G2:H2')  # slip no data
                        sheet.merge_cells('A3:F3')  # 'HEADING' 
                        sheet.merge_cells('A4:F5')  # heading data
                        sheet.merge_cells('A6:F6')  # 'DESCRIPTION'
                        sheet.merge_cells('A7:F8')  # description data
                        sheet.merge_cells('G2:H2')  # slipno data
                        sheet.merge_cells('G3:H3')  # 'UNIT'
                        sheet.merge_cells('G4:H5')  # unit data
                        sheet.merge_cells('G5:H5')  # empty
                        sheet.merge_cells('G6:H6')  # 'QUANTITY'
                        sheet.merge_cells('G7:H7')  # qty data
                        sheet.merge_cells('G8:H8')  # 'TAKER OFF'
                        sheet.merge_cells('G9:H9')  # Taker off data
                        sheet.merge_cells('G10:H10')  # 'SQD'
                        sheet.merge_cells('G11:H11')    # sqd data
                        sheet.merge_cells('G12:H12')    # 'CH.SQD'
                        sheet.merge_cells('G13:H13')    # ch sqd data
                        sheet.merge_cells('G15:H15')    # red data
                        sheet.merge_cells('G14:H14')    # 'RED'
                        sheet.merge_cells('G16:H16')    # 'RATE'
                        sheet.merge_cells('G17:H17')    # rate data
                        sheet.merge_cells('G18:H18')    # 'TOTAL COST'
                        sheet.merge_cells('G19:H19')    # total cost data
                        sheet.merge_cells('G20:H20')    # 'DATE AND TIME'
                        sheet.merge_cells('G21:H21')    # date & time
                        sheet.merge_cells('G22:H22')    # empty
                        sheet.merge_cells('G23:H23')    # empty
                        sheet.merge_cells('D9:F9')    # empty

                        sheet["A1"].alignment = Alignment(horizontal="center")
                        sheet['C1'].alignment = Alignment(horizontal="center")
                        sheet['E1'].alignment = Alignment(horizontal="center")
                        sheet['G1'].alignment = Alignment(horizontal="center")
                        sheet['A3'].alignment = Alignment(horizontal="center")
                        sheet['A4'].alignment = Alignment(horizontal="center", vertical="center")
                        sheet['G3'].alignment = Alignment(horizontal="center")
                        sheet['G5'].alignment = Alignment(horizontal="center")
                        sheet['A6'].alignment = Alignment(horizontal="center")
                        sheet['A7'].alignment = Alignment(horizontal="center", vertical="center")
                        sheet['G7'].alignment = Alignment(horizontal="center")
                        sheet['G9'].alignment = Alignment(horizontal="center")
                        sheet['G11'].alignment = Alignment(horizontal="center")
                        sheet['G13'].alignment = Alignment(horizontal="center")
                        sheet['G15'].alignment = Alignment(horizontal="center")
                        sheet['G17'].alignment = Alignment(horizontal="center")
                        sheet['G19'].alignment = Alignment(horizontal="center")
                        sheet["A2"].alignment = Alignment(horizontal="center")
                        sheet['C2'].alignment = Alignment(horizontal="center")
                        sheet['E2'].alignment = Alignment(horizontal="center")
                        sheet['G2'].alignment = Alignment(horizontal="center")
                        sheet['G4'].alignment = Alignment(horizontal="center", vertical="center")
                        sheet['G6'].alignment = Alignment(horizontal="center")
                        sheet['G8'].alignment = Alignment(horizontal="center")
                        sheet['G10'].alignment = Alignment(horizontal="center")
                        sheet['G12'].alignment = Alignment(horizontal="center")
                        sheet['G14'].alignment = Alignment(horizontal="center")
                        sheet['G16'].alignment = Alignment(horizontal="center")
                        sheet['G18'].alignment = Alignment(horizontal="center")
                        sheet['G20'].alignment = Alignment(horizontal="center")
                        sheet['G21'].alignment = Alignment(horizontal="center")
                        sheet['G22'].alignment = Alignment(horizontal="center")
                        sheet['G23'].alignment = Alignment(horizontal="center")

                        sheet.column_dimensions['A'].width = 13
                        sheet.column_dimensions['B'].width = 13
                        sheet.column_dimensions['C'].width = 15
                        sheet.column_dimensions['D'].width = 20
                        sheet.column_dimensions['E'].width = 3.5
                        sheet.column_dimensions['F'].width = 22
                        sheet.column_dimensions['G'].width = 15
                        sheet.column_dimensions['H'].width = 15
                        sheet.column_dimensions['I'].width = 15
                        sheet.column_dimensions['J'].width = 15
                        sheet.column_dimensions['K'].width = 10
                        sheet.column_dimensions['L'].width = 10
                        sheet.column_dimensions['M'].width = 14
                        sheet.column_dimensions['N'].width = 30
                        sheet.cell(row=1, column=1).font = Font(bold=True)
                        sheet.cell(row=1, column=3).font = Font(bold=True)
                        sheet.cell(row=1, column=5).font = Font(bold=True)
                        sheet.cell(row=1, column=7).font = Font(bold=True)
                        sheet.cell(row=3, column=1).font = Font(bold=True)
                        sheet.cell(row=3, column=7).font = Font(bold=True)
                        sheet.cell(row=6, column=1).font = Font(bold=True)
                        sheet.cell(row=6, column=7).font = Font(bold=True)
                        sheet.cell(row=8, column=7).font = Font(bold=True)
                        sheet.cell(row=10, column=7).font = Font(bold=True)
                        sheet.cell(row=12, column=7).font = Font(bold=True)
                        sheet.cell(row=14, column=7).font = Font(bold=True)
                        sheet.cell(row=16, column=7).font = Font(bold=True)
                        sheet.cell(row=18, column=7).font = Font(bold=True)
                        sheet.cell(row=20, column=7).font = Font(bold=True)
                        sheet.cell(row=22, column=7).font = Font(bold=True)
                        sheet.cell(row=1, column=7).value = "SLIP NO :"
                        sheet.cell(row=3, column=1).value = "HEADING :"
                        sheet.cell(row=6, column=1).value = "DESCRIPTION :"
                        sheet.cell(row=3, column=7).value = "UNIT :"
                        sheet.cell(row=6, column=7).value = "QUANTITY :"
                        sheet.cell(row=16, column=7).value = "RATE :"
                        sheet.cell(row=18, column=7).value = "TOTAL COST :"
                        sheet.cell(row=10, column=7).value = "SQD :"
                        sheet.cell(row=12, column=7).value = "CH.SQD :"
                        sheet.cell(row=14, column=7).value = "RED :"
                        sheet.cell(row=1, column=3).value = "BIL NO :"
                        sheet.cell(row=1, column=1).value = "JOB :"
                        sheet.cell(row=1, column=5).value = "ELEMENT NO :"
                        sheet.cell(row=8, column=7).value = "TAKER OFF :"
                        sheet.cell(row=20, column=7).value = "DATE AND TIME :"
                        sheet.cell(row=2, column=7).value = slip_no
                        sheet.cell(row=4, column=1).value = heading
                        sheet.cell(row=7, column=1).value = description
                        sheet.cell(row=4, column=7).value = unit
                        sheet.cell(row=7, column=7).value = quantity
                        sheet.cell(row=17, column=7).value = cost
                        sheet.cell(row=19, column=7).value = total_cost
                        sheet.cell(row=11, column=7).value = sqd
                        sheet.cell(row=13, column=7).value = ch_sqd
                        sheet.cell(row=15, column=7).value = red
                        sheet.cell(row=2, column=3).value = bill
                        sheet.cell(row=2, column=1).value = project_name
                        sheet.cell(row=2, column=5).value = element
                        sheet.cell(row=9, column=7).value = taker
                        sheet.cell(row=21, column=7).value = current_time
                        
                        # apply borders
                        # row 1
                        sheet.cell(row=1, column=1).border = Border(left=Side(style='thick'), top=Side(style='thick'))
                        sheet.cell(row=1, column=2).border = Border(right=Side(style='thick'), top=Side(style='thick'))
                        sheet.cell(row=1, column=3).border = Border(left=Side(style='thick'), top=Side(style='thick'))
                        sheet.cell(row=1, column=4).border = Border(right=Side(style='thick'), top=Side(style='thick'))
                        sheet.cell(row=1, column=5).border = Border(left=Side(style='thick'), top=Side(style='thick'))
                        sheet.cell(row=1, column=6).border = Border(right=Side(style='thick'), top=Side(style='thick'))
                        sheet.cell(row=1, column=7).border = Border(left=Side(style='thick'), top=Side(style='thick'))
                        sheet.cell(row=1, column=8).border = Border(right=Side(style='thick'), top=Side(style='thick'))
                        
                        # row 2
                        sheet.cell(row=2, column=1).border = Border(left=Side(style='thick'))
                        sheet.cell(row=2, column=2).border = Border(right=Side(style='thick'))
                        sheet.cell(row=2, column=4).border = Border(right=Side(style="thick"))
                        sheet.cell(row=2, column=6).border = Border(right=Side(style='thick'))
                        sheet.cell(row=2, column=8).border = Border(right=Side(style='thick'))
                        
                        # row 3
                        sheet.cell(row=3, column=1).border = Border(left=Side(style="thick"), right=Side(style="thick"), top=Side(style="thick"))
                        sheet.cell(row=3, column=2).border = Border(top=Side(style="thick"))
                        sheet.cell(row=3, column=3).border = Border(top=Side(style="thick"))
                        sheet.cell(row=3, column=4).border = Border(top=Side(style="thick"))
                        sheet.cell(row=3, column=5).border = Border(top=Side(style="thick"))
                        sheet.cell(row=3, column=6).border = Border(right=Side(style="thick"), top=Side(style="thick"))
                        sheet.cell(row=3, column=7).border = Border(left=Side(style="thick"), top=Side(style="thick"))
                        sheet.cell(row=3, column=8).border = Border(right=Side(style="thick"), top=Side(style="thick"))
                        
                        # row 4
                        sheet.cell(row=4, column=1).border = Border(left=Side(style="thick"))
                        sheet.cell(row=4, column=6).border = Border(right=Side(style="thick"))
                        sheet.cell(row=4, column=7).border = Border(left=Side(style="thick"))
                        sheet.cell(row=4, column=8).border = Border(right=Side(style="thick"))
                        
                        # row 5
                        sheet.cell(row=5, column=1).border = Border(left=Side(style="thick"))
                        sheet.cell(row=5, column=6).border = Border(right=Side(style="thick"))
                        sheet.cell(row=5, column=8).border = Border(right=Side(style="thick"))

                        # row 6
                        sheet.cell(row=6, column=1).border = Border(left=Side(style="thick"), top=Side(style="thick"))
                        sheet.cell(row=6, column=2).border = Border(top=Side(style="thick"))
                        sheet.cell(row=6, column=3).border = Border(top=Side(style="thick"))
                        sheet.cell(row=6, column=4).border = Border(top=Side(style="thick"))
                        sheet.cell(row=6, column=5).border = Border(top=Side(style="thick"))
                        sheet.cell(row=6, column=6).border = Border(right=Side(style="thick"), top=Side(style="thick"))
                        sheet.cell(row=6, column=7).border = Border(top=Side(style="thick"))
                        sheet.cell(row=6, column=8).border = Border(right=Side(style="thick"), top=Side(style="thick"))

                        # row 7
                        sheet.cell(row=7, column=1).border = Border(left=Side(style="thick"))
                        sheet.cell(row=7, column=6).border = Border(right=Side(style="thick"))
                        sheet.cell(row=7, column=7).border = Border(bottom=Side(style="thick"))
                        sheet.cell(row=7, column=8).border = Border(right=Side(style="thick"), bottom=Side(style="thick"))
                        
                        # row 8
                        sheet.cell(row=8, column=1).border = Border(left=Side(style="thick"), bottom=Side(style="thick"))
                        sheet.cell(row=8, column=2).border = Border(bottom=Side(style="thick"))
                        sheet.cell(row=8, column=3).border = Border(bottom=Side(style="thick"))
                        sheet.cell(row=8, column=4).border = Border(bottom=Side(style="thick"))
                        sheet.cell(row=8, column=5).border = Border(bottom=Side(style="thick"))
                        sheet.cell(row=8, column=6).border = Border(right=Side(style="thick"), bottom=Side(style="thick"))
                        sheet.cell(row=8, column=7).border = Border(left=Side(style="thick"), top=Side(style="thick"))
                        sheet.cell(row=8, column=8).border = Border(right=Side(style="thick"), top=Side(style="thick"))

                        # row 9
                        sheet.cell(row=9, column=1).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=9, column=2).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=9, column=3).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=9, column=7).border = Border(left=Side(style="thick"), bottom=Side(style="thick"))
                        sheet.cell(row=9, column=8).border = Border(right=Side(style="thick"), bottom=Side(style="thick"))
                        
                        # row 10
                        sheet.cell(row=10, column=1).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=10, column=2).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=10, column=3).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=10, column=7).border = Border(left=Side(style="thick"))
                        sheet.cell(row=10, column=8).border = Border(right=Side(style="thick"))
                        
                        # row 11
                        sheet.cell(row=11, column=1).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=11, column=2).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=11, column=3).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=11, column=7).border = Border(left=Side(style="thick"), bottom=Side(style="thick"))
                        sheet.cell(row=11, column=8).border = Border(right=Side(style="thick"), bottom=Side(style="thick"))
                        
                        # row 12
                        sheet.cell(row=12, column=1).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=12, column=2).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=12, column=3).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=12, column=7).border = Border(left=Side(style="thick"))
                        sheet.cell(row=12, column=8).border = Border(right=Side(style="thick"))
                        
                        # row 13
                        sheet.cell(row=13, column=1).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=13, column=2).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=13, column=3).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=13, column=7).border = Border(left=Side(style="thick"), bottom=Side(style="thick"))
                        sheet.cell(row=13, column=8).border = Border(right=Side(style="thick"), bottom=Side(style="thick"))
                        
                        
                        # row 14
                        sheet.cell(row=14, column=1).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=14, column=2).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=14, column=3).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=14, column=7).border = Border(left=Side(style="thick"))
                        sheet.cell(row=14, column=8).border = Border(right=Side(style="thick"))
                        
                        # row 15
                        sheet.cell(row=15, column=1).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=15, column=2).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=15, column=3).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=15, column=7).border = Border(left=Side(style="thick"), bottom=Side(style="thick"))
                        sheet.cell(row=15, column=8).border = Border(right=Side(style="thick"), bottom=Side(style="thick"))
                        
                        
                        # row 16
                        sheet.cell(row=16, column=1).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=16, column=2).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=16, column=3).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=16, column=7).border = Border(left=Side(style="thick"))
                        sheet.cell(row=16, column=8).border = Border(right=Side(style="thick"))
                        
                        # row 17
                        sheet.cell(row=17, column=1).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=17, column=2).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=17, column=3).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=17, column=7).border = Border(left=Side(style="thick"), bottom=Side(style="thick"))
                        sheet.cell(row=17, column=8).border = Border(right=Side(style="thick"), bottom=Side(style="thick"))                        
                        
                        
                        # row 18
                        sheet.cell(row=18, column=1).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=18, column=2).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=18, column=3).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=18, column=7).border = Border(left=Side(style="thick"))
                        sheet.cell(row=18, column=8).border = Border(right=Side(style="thick"))
                        
                        # row 19
                        sheet.cell(row=19, column=1).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=19, column=2).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=19, column=3).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=19, column=7).border = Border(left=Side(style="thick"), bottom=Side(style="thick"))
                        sheet.cell(row=19, column=8).border = Border(right=Side(style="thick"), bottom=Side(style="thick"))                        
                                                
                        # row 20
                        sheet.cell(row=20, column=1).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=20, column=2).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=20, column=3).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=20, column=7).border = Border(left=Side(style="thick"))
                        sheet.cell(row=20, column=8).border = Border(right=Side(style="thick"))
                        
                        # row 21
                        sheet.cell(row=21, column=1).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=21, column=2).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=21, column=3).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=21, column=7).border = Border(left=Side(style="thick"), bottom=Side(style="thick"))
                        sheet.cell(row=21, column=8).border = Border(right=Side(style="thick"), bottom=Side(style="thick"))    
                        
                        # row 22
                        sheet.cell(row=22, column=1).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=22, column=2).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=22, column=3).border = Border(left=Side(style="thick"), right=Side(style="thick"))            
                        
                        # row 23
                        sheet.cell(row=23, column=1).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=23, column=2).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=23, column=3).border = Border(left=Side(style="thick"), right=Side(style="thick"))               
                        
                        
                        # row 24
                        sheet.cell(row=24, column=1).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=24, column=2).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=24, column=3).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        
                        
                        # row 25
                        sheet.cell(row=25, column=1).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=25, column=2).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=25, column=3).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        
                        # row 26
                        sheet.cell(row=26, column=1).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=26, column=2).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=26, column=3).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                                                
                        # row 27
                        sheet.cell(row=27, column=1).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=27, column=2).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=27, column=3).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                                                
                        # row 28
                        sheet.cell(row=28, column=1).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=28, column=2).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=28, column=3).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                                                
                        # row 29
                        sheet.cell(row=29, column=1).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=29, column=2).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=29, column=3).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                                                
                        # row 30
                        sheet.cell(row=30, column=1).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=30, column=2).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=30, column=3).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                                                
                        # row 31
                        sheet.cell(row=31, column=1).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=31, column=2).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=31, column=3).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                                                                        
                        # row 32
                        sheet.cell(row=32, column=1).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=32, column=2).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=32, column=3).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                                                                        
                        # row 33
                        sheet.cell(row=33, column=1).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=33, column=2).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=33, column=3).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                                                                        
                        # row 34
                        sheet.cell(row=34, column=1).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=34, column=2).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        sheet.cell(row=34, column=3).border = Border(left=Side(style="thick"), right=Side(style="thick"))
                        
                        img = Image.open("my_image.png")
                        img = openpyxl.drawing.image.Image("my_image.png")
                        sheet.add_image(img, "D10")

                        if unit_box_cb.get() == "m" or unit_box_cb.get() == 'no':

                            sheet.cell(row=9, column=1).value = c1r1
                            sheet.cell(row=11, column=1).value = c1r3
                            sheet.cell(row=13, column=1).value = c1r5
                            sheet.cell(row=15, column=1).value = c1r7
                            sheet.cell(row=17, column=1).value = c1r9
                            sheet.cell(row=19, column=1).value = c1r11
                            sheet.cell(row=21, column=1).value = c1r13
                            sheet.cell(row=23, column=1).value = c1r15
                            sheet.cell(row=25, column=1).value = c1r17
                            sheet.cell(row=27, column=1).value = c1r19
                            sheet.cell(row=29, column=1).value = c1r21
                            sheet.cell(row=31, column=1).value = c1r23
                            sheet.cell(row=9, column=2).value = c2r1
                            sheet.cell(row=11, column=2).value = c2r3
                            sheet.cell(row=13, column=2).value = c2r5
                            sheet.cell(row=15, column=2).value = c2r7
                            sheet.cell(row=17, column=2).value = c2r9
                            sheet.cell(row=19, column=2).value = c2r11
                            sheet.cell(row=21, column=2).value = c2r13
                            sheet.cell(row=23, column=2).value = c2r15
                            sheet.cell(row=25, column=2).value = c2r17
                            sheet.cell(row=27, column=2).value = c2r19
                            sheet.cell(row=29, column=2).value = c2r21
                            sheet.cell(row=31, column=2).value = c2r23
                            sheet.cell(row=10, column=3).value = c3r2
                            sheet.cell(row=12, column=3).value = c3r4
                            sheet.cell(row=14, column=3).value = c3r6
                            sheet.cell(row=16, column=3).value = c3r8
                            sheet.cell(row=18, column=3).value = c3r10
                            sheet.cell(row=20, column=3).value = c3r12
                            sheet.cell(row=22, column=3).value = c3r14
                            sheet.cell(row=24, column=3).value = c3r16
                            sheet.cell(row=26, column=3).value = c3r18
                            sheet.cell(row=28, column=3).value = c3r20
                            sheet.cell(row=30, column=3).value = c3r22
                            sheet.cell(row=32, column=3).value = c3r24
                            sheet.cell(row=33, column=3).value = c3r25

                        elif unit_box_cb.get() == "kg":

                            sheet.cell(row=9, column=1).value = c1r1
                            sheet.cell(row=11, column=1).value = c1r3
                            sheet.cell(row=13, column=1).value = c1r5
                            sheet.cell(row=15, column=1).value = c1r7
                            sheet.cell(row=17, column=1).value = c1r9
                            sheet.cell(row=19, column=1).value = c1r11
                            sheet.cell(row=21, column=1).value = c1r13
                            sheet.cell(row=23, column=1).value = c1r15
                            sheet.cell(row=25, column=1).value = c1r17
                            sheet.cell(row=27, column=1).value = c1r19
                            sheet.cell(row=29, column=1).value = c1r21
                            sheet.cell(row=31, column=1).value = c1r23
                            sheet.cell(row=9, column=2).value = c2r1
                            sheet.cell(row=11, column=2).value = c2r3
                            sheet.cell(row=13, column=2).value = c2r5
                            sheet.cell(row=15, column=2).value = c2r7
                            sheet.cell(row=17, column=2).value = c2r9
                            sheet.cell(row=19, column=2).value = c2r11
                            sheet.cell(row=21, column=2).value = c2r13
                            sheet.cell(row=23, column=2).value = c2r15
                            sheet.cell(row=25, column=2).value = c2r17
                            sheet.cell(row=27, column=2).value = c2r19
                            sheet.cell(row=29, column=2).value = c2r21
                            sheet.cell(row=31, column=2).value = c2r23
                            sheet.cell(row=10, column=3).value = c3r2
                            sheet.cell(row=12, column=3).value = c3r4
                            sheet.cell(row=14, column=3).value = c3r6
                            sheet.cell(row=16, column=3).value = c3r8
                            sheet.cell(row=18, column=3).value = c3r10
                            sheet.cell(row=20, column=3).value = c3r12
                            sheet.cell(row=22, column=3).value = c3r14
                            sheet.cell(row=24, column=3).value = c3r16
                            sheet.cell(row=26, column=3).value = c3r18
                            sheet.cell(row=28, column=3).value = c3r20
                            sheet.cell(row=30, column=3).value = c3r22
                            sheet.cell(row=32, column=3).value = c3r24
                            sheet.cell(row=33, column=3).value = c3r25
                            sheet.cell(row=31, column=4).value = kg
                            sheet.cell(row=31, column=5).value = "X"
                            sheet.cell(row=31, column=6).value = kg_new + "kg/m"

                        elif unit_box_cb.get() == "item":
                            sheet.cell(row=17, column=2).value = "ITEM"

                        elif unit_box_cb.get() == "m2":
                            sheet.cell(row=9, column=1).value = c1r1
                            sheet.cell(row=12, column=1).value = c1r4
                            sheet.cell(row=15, column=1).value = c1r7
                            sheet.cell(row=18, column=1).value = c1r10
                            sheet.cell(row=21, column=1).value = c1r13
                            sheet.cell(row=24, column=1).value = c1r16
                            sheet.cell(row=27, column=1).value = c1r19
                            sheet.cell(row=30, column=1).value = c1r22
                            sheet.cell(row=9, column=2).value = c2r1
                            sheet.cell(row=10, column=2).value = c2r2
                            sheet.cell(row=12, column=2).value = c2r4
                            sheet.cell(row=13, column=2).value = c2r5
                            sheet.cell(row=15, column=2).value = c2r7
                            sheet.cell(row=16, column=2).value = c2r8
                            sheet.cell(row=18, column=2).value = c2r10
                            sheet.cell(row=19, column=2).value = c2r11
                            sheet.cell(row=21, column=2).value = c2r13
                            sheet.cell(row=22, column=2).value = c2r14
                            sheet.cell(row=24, column=2).value = c2r16
                            sheet.cell(row=25, column=2).value = c2r17
                            sheet.cell(row=27, column=2).value = c2r19
                            sheet.cell(row=28, column=2).value = c2r20
                            sheet.cell(row=30, column=2).value = c2r22
                            sheet.cell(row=31, column=2).value = c2r23
                            sheet.cell(row=11, column=3).value = c3r3
                            sheet.cell(row=14, column=3).value = c3r6
                            sheet.cell(row=17, column=3).value = c3r9
                            sheet.cell(row=20, column=3).value = c3r12
                            sheet.cell(row=23, column=3).value = c3r15
                            sheet.cell(row=26, column=3).value = c3r18
                            sheet.cell(row=29, column=3).value = c3r21
                            sheet.cell(row=32, column=3).value = c3r24
                            sheet.cell(row=33, column=3).value = c3r25

                        elif unit_box_cb.get() == "m3":
                            sheet.cell(row=9, column=1).value = c1r1
                            sheet.cell(row=13, column=1).value = c1r5
                            sheet.cell(row=17, column=1).value = c1r9
                            sheet.cell(row=21, column=1).value = c1r13
                            sheet.cell(row=25, column=1).value = c1r17
                            sheet.cell(row=29, column=1).value = c1r21
                            sheet.cell(row=9, column=2).value = c2r1
                            sheet.cell(row=10, column=2).value = c2r2
                            sheet.cell(row=11, column=2).value = c2r3
                            sheet.cell(row=13, column=2).value = c2r5
                            sheet.cell(row=14, column=2).value = c2r6
                            sheet.cell(row=15, column=2).value = c2r7
                            sheet.cell(row=17, column=2).value = c2r9
                            sheet.cell(row=18, column=2).value = c2r10
                            sheet.cell(row=19, column=2).value = c2r11
                            sheet.cell(row=21, column=2).value = c2r13
                            sheet.cell(row=22, column=2).value = c2r14
                            sheet.cell(row=23, column=2).value = c2r15
                            sheet.cell(row=25, column=2).value = c2r17
                            sheet.cell(row=26, column=2).value = c2r18
                            sheet.cell(row=27, column=2).value = c2r19
                            sheet.cell(row=29, column=2).value = c2r21
                            sheet.cell(row=30, column=2).value = c2r22
                            sheet.cell(row=31, column=2).value = c2r23
                            sheet.cell(row=12, column=3).value = c3r4
                            sheet.cell(row=16, column=3).value = c3r8
                            sheet.cell(row=20, column=3).value = c3r12
                            sheet.cell(row=24, column=3).value = c3r16
                            sheet.cell(row=28, column=3).value = c3r20
                            sheet.cell(row=32, column=3).value = c3r24
                            sheet.cell(row=33, column=3).value = c3r25

                        else:
                            pass

                        file.save("TAKING_OFF.xlsx")

                        tko.destroy()
                        root.deiconify()


                def back():
                    tko.destroy()
                    root.deiconify()

                tko = Tk()
                tko.title('TAKING OFF PAPER')
                tko.geometry("800x600+280+30")

                tko.rowconfigure(0, weight=1)
                tko.columnconfigure(0, weight=1)

                menubar = Menu(tko)
                tko.config(menu=menubar)

                # pack to screen
                # calculation frame
                calculation_frame = Frame(tko)
                calculation_frame.pack_forget()

                # convert frame
                covert_frame = Frame(tko)
                covert_frame.place_forget()

                # taker off frame
                taker_frame = Frame(tko)
                taker_frame.place(x=615, y=127)

                taker_off_label = Label(taker_frame, text="TAKER OFF")
                taker_off_label.grid(row=0, column=0)

                sqd_label = Label(taker_frame, text="SQD")
                sqd_label.grid(row=2, column=0)

                ch_sqd_label = Label(taker_frame, text="CH.SQD")
                ch_sqd_label.grid(row=4, column=0)

                red_label = Label(taker_frame, text="RED")
                red_label.grid(row=6, column=0)

                # rate_label = Label(taker_frame, text="RATE", fg="#0000FF", bg="yellow")
                rate_label = Label(taker_frame, text="RATE")
                rate_label.grid(row=8, column=0)

                total_cost_label = Label(taker_frame, text="TOTAL COST")
                total_cost_label.grid(row=10, column=0)

                date_time_label = Label(taker_frame, text="DATE AND TIME")
                date_time_label.grid(row=12, column=0)

                job_frame = Frame(tko)
                job_frame.place(x=10, y=5)

                jl = Label(job_frame, text="JOB")
                jl.grid(row=0, column=0)

                bill = Label(job_frame, text="BILL NO.")
                bill.grid(row=0, column=1)

                element = Label(job_frame, text="ELEMENT NO.")
                element.grid(row=0, column=2)

                slip = Label(job_frame, text="SLIP NO.")
                slip.grid(row=0, column=3)

                heading_frame = Frame(tko)
                heading_frame.place(x=10, y=45)

                hl = Label(heading_frame, text="HEADING")
                hl.grid(row=0, column=0)

                dl = Label(heading_frame, text="DESCRIPTION")
                dl.grid(row=2, column=0)

                # ul = Label(heading_frame, text="UNIT", fg="#0000FF", bg="yellow")
                ul = Label(heading_frame, text="UNIT")
                ul.grid(row=0, column=3)

                ql = Label(heading_frame, text="QUANTITY")
                ql.grid(row=2, column=3)

                item_label = Label(tko, text="ITEM")
                item_label.pack_forget()

                time_current_label = Label(taker_frame, text=current_time)
                time_current_label.grid(row=13, column=0)

                # ENTRY BOX
                heading_box = Entry((heading_frame), width=100)
                heading_box.grid(row=1, column=0)

                description_box = Entry((heading_frame), width=100)
                description_box.grid(row=3, column=0)

                quantity_paper_box = Entry(heading_frame, width=27, justify='center')
                quantity_paper_box.grid(row=3, column=3)

                global unit_box_cb
                n = StringVar()
                unit_box_cb = ttk.Combobox(heading_frame, width=25, textvariable=n, justify='center')
                unit_box_cb.config(values=('m', 'm2', 'm3', 'kg', 'item', 'no'))
                unit_box_cb.set('Please Select One')
                unit_box_cb.grid(row=1, column=3)

                job_box = Entry((job_frame), width=33, justify='center')
                job_box.grid(row=1, column=0)
                job_box.insert(0, project_box.get())
                job_box.configure(state=DISABLED)

                billTextVariable = StringVar()
                bill_box = ttk.Combobox(job_frame, width=28, textvariable=billTextVariable, justify='center')
                bill_box.config(values=('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'))
                # bill_box = Entry((job_frame), width=32, justify='center')
                bill_box.grid(row=1, column=1)

                elementTextVariable = StringVar()
                element_box = ttk.Combobox(job_frame, width=28, textvariable=elementTextVariable, justify='center')
                element_box.config(values=('1', '2', '3', '4', '5', '6', '7', '8', '9', '10'))
                # element_box = Entry(job_frame, width=35, justify='center')
                element_box.grid(row=1, column=2)


                slipTextVariable = StringVar()
                slip_box = ttk.Combobox(job_frame, width=28, textvariable=slipTextVariable, justify='center')
                slip_box.config(values=('1', '2', '3', '4', '5', '6', '7', '8', '9', '10'))
                # slip_box = Entry(job_frame, width=27, justify='center')
                slip_box.grid(row=1, column=3)

                taker_off_box = Entry((taker_frame), width=28, justify='center')
                taker_off_box.grid(row=1, column=0)
                taker_off_box.insert(0, taker_off_name_box.get())
                taker_off_box.configure(state=DISABLED)

                sqd_box = Entry((taker_frame), width=28, justify='center')
                sqd_box.grid(row=3, column=0)

                ch_sqd_box = Entry(taker_frame, width=28, justify='center')
                ch_sqd_box.grid(row=5, column=0)

                red_box = Entry(taker_frame, width=28, justify='center')
                red_box.grid(row=7, column=0)

                rate_box = Entry(taker_frame, width=28, justify='center')
                rate_box.grid(row=9, column=0)

                total_cost_box = Entry(taker_frame, width=28, justify='center')
                total_cost_box.grid(row=11, column=0)

                # calculation frame entry box

                r1c1 = Entry(calculation_frame, width=15, justify='center')
                r1c1.grid(row=0, column=0)
                r2c1 = Entry(calculation_frame, width=15, justify='center')
                r2c1.grid(row=1, column=0)
                r3c1 = Entry(calculation_frame, width=15, justify='center')
                r3c1.grid(row=2, column=0)
                r4c1 = Entry(calculation_frame, width=15, justify='center')
                r4c1.grid(row=3, column=0)
                r5c1 = Entry(calculation_frame, width=15, justify='center')
                r5c1.grid(row=4, column=0)
                r6c1 = Entry(calculation_frame, width=15, justify='center')
                r6c1.grid(row=5, column=0)
                r7c1 = Entry(calculation_frame, width=15, justify='center')
                r7c1.grid(row=6, column=0)
                r8c1 = Entry(calculation_frame, width=15, justify='center')
                r8c1.grid(row=7, column=0)
                r9c1 = Entry(calculation_frame, width=15, justify='center')
                r9c1.grid(row=8, column=0)
                r10c1 = Entry(calculation_frame, width=15, justify='center')
                r10c1.grid(row=9, column=0)
                r11c1 = Entry(calculation_frame, width=15, justify='center')
                r11c1.grid(row=10, column=0)
                r12c1 = Entry(calculation_frame, width=15, justify='center')
                r12c1.grid(row=11, column=0)
                r13c1 = Entry(calculation_frame, width=15, justify='center')
                r13c1.grid(row=12, column=0)
                r14c1 = Entry(calculation_frame, width=15, justify='center')
                r14c1.grid(row=13, column=0)
                r15c1 = Entry(calculation_frame, width=15, justify='center')
                r15c1.grid(row=14, column=0)
                r16c1 = Entry(calculation_frame, width=15, justify='center')
                r16c1.grid(row=15, column=0)
                r17c1 = Entry(calculation_frame, width=15, justify='center')
                r17c1.grid(row=16, column=0)
                r18c1 = Entry(calculation_frame, width=15, justify='center')
                r18c1.grid(row=17, column=0)
                r19c1 = Entry(calculation_frame, width=15, justify='center')
                r19c1.grid(row=18, column=0)
                r20c1 = Entry(calculation_frame, width=15, justify='center')
                r20c1.grid(row=19, column=0)
                r21c1 = Entry(calculation_frame, width=15, justify='center')
                r21c1.grid(row=20, column=0)
                r22c1 = Entry(calculation_frame, width=15, justify='center')
                r22c1.grid(row=21, column=0)
                r23c1 = Entry(calculation_frame, width=15, justify='center')
                r23c1.grid(row=22, column=0)
                r24c1 = Entry(calculation_frame, width=15, justify='center')
                r24c1.grid(row=23, column=0)
                r1c2 = Entry(calculation_frame, width=15, justify='center')
                r1c2.grid(row=0, column=1)
                r2c2 = Entry(calculation_frame, width=15, justify='center')
                r2c2.grid(row=1, column=1)
                r3c2 = Entry(calculation_frame, width=15, justify='center')
                r3c2.grid(row=2, column=1)
                r4c2 = Entry(calculation_frame, width=15, justify='center')
                r4c2.grid(row=3, column=1)
                r5c2 = Entry(calculation_frame, width=15, justify='center')
                r5c2.grid(row=4, column=1)
                r6c2 = Entry(calculation_frame, width=15, justify='center')
                r6c2.grid(row=5, column=1)
                r7c2 = Entry(calculation_frame, width=15, justify='center')
                r7c2.grid(row=6, column=1)
                r8c2 = Entry(calculation_frame, width=15, justify='center')
                r8c2.grid(row=7, column=1)
                r9c2 = Entry(calculation_frame, width=15, justify='center')
                r9c2.grid(row=8, column=1)
                r10c2 = Entry(calculation_frame, width=15, justify='center')
                r10c2.grid(row=9, column=1)
                r11c2 = Entry(calculation_frame, width=15, justify='center')
                r11c2.grid(row=10, column=1)
                r12c2 = Entry(calculation_frame, width=15, justify='center')
                r12c2.grid(row=11, column=1)
                r13c2 = Entry(calculation_frame, width=15, justify='center')
                r13c2.grid(row=12, column=1)
                r14c2 = Entry(calculation_frame, width=15, justify='center')
                r14c2.grid(row=13, column=1)
                r15c2 = Entry(calculation_frame, width=15, justify='center')
                r15c2.grid(row=14, column=1)
                r16c2 = Entry(calculation_frame, width=15, justify='center')
                r16c2.grid(row=15, column=1)
                r17c2 = Entry(calculation_frame, width=15, justify='center')
                r17c2.grid(row=16, column=1)
                r18c2 = Entry(calculation_frame, width=15, justify='center')
                r18c2.grid(row=17, column=1)
                r19c2 = Entry(calculation_frame, width=15, justify='center')
                r19c2.grid(row=18, column=1)
                r20c2 = Entry(calculation_frame, width=15, justify='center')
                r20c2.grid(row=19, column=1)
                r21c2 = Entry(calculation_frame, width=15, justify='center')
                r21c2.grid(row=20, column=1)
                r22c2 = Entry(calculation_frame, width=15, justify='center')
                r22c2.grid(row=21, column=1)
                r23c2 = Entry(calculation_frame, width=15, justify='center')
                r23c2.grid(row=22, column=1)
                r24c2 = Entry(calculation_frame, width=15, justify='center')
                r24c2.grid(row=23, column=1)

                r1c3 = Entry(calculation_frame, width=15, justify='center', bg='yellow')
                r1c3.grid(row=0, column=2)
                r2c3 = Entry(calculation_frame, width=15, justify='center', bg='yellow')
                r2c3.grid(row=1, column=2)
                r3c3 = Entry(calculation_frame, width=15, justify='center', bg='yellow')
                r3c3.grid(row=2, column=2)
                r4c3 = Entry(calculation_frame, width=15, justify='center', bg='yellow')
                r4c3.grid(row=3, column=2)
                r5c3 = Entry(calculation_frame, width=15, justify='center', bg='yellow')
                r5c3.grid(row=4, column=2)
                r6c3 = Entry(calculation_frame, width=15, justify='center', bg='yellow')
                r6c3.grid(row=5, column=2)
                r7c3 = Entry(calculation_frame, width=15, justify='center', bg='yellow')
                r7c3.grid(row=6, column=2)
                r8c3 = Entry(calculation_frame, width=15, justify='center', bg='yellow')
                r8c3.grid(row=7, column=2)
                r9c3 = Entry(calculation_frame, width=15, justify='center', bg='yellow')
                r9c3.grid(row=8, column=2)
                r10c3 = Entry(calculation_frame, width=15, justify='center', bg='yellow')
                r10c3.grid(row=9, column=2)
                r11c3 = Entry(calculation_frame, width=15, justify='center', bg='yellow')
                r11c3.grid(row=10, column=2)
                r12c3 = Entry(calculation_frame, width=15, justify='center', bg='yellow')
                r12c3.grid(row=11, column=2)
                r13c3 = Entry(calculation_frame, width=15, justify='center', bg='yellow')
                r13c3.grid(row=12, column=2)
                r14c3 = Entry(calculation_frame, width=15, justify='center', bg='yellow')
                r14c3.grid(row=13, column=2)
                r15c3 = Entry(calculation_frame, width=15, justify='center', bg='yellow')
                r15c3.grid(row=14, column=2)
                r16c3 = Entry(calculation_frame, width=15, justify='center', bg='yellow')
                r16c3.grid(row=15, column=2)
                r17c3 = Entry(calculation_frame, width=15, justify='center', bg='yellow')
                r17c3.grid(row=16, column=2)
                r18c3 = Entry(calculation_frame, width=15, justify='center', bg='yellow')
                r18c3.grid(row=17, column=2)
                r19c3 = Entry(calculation_frame, width=15, justify='center', bg='yellow')
                r19c3.grid(row=18, column=2)
                r20c3 = Entry(calculation_frame, width=15, justify='center', bg='yellow')
                r20c3.grid(row=19, column=2)
                r21c3 = Entry(calculation_frame, width=15, justify='center', bg='yellow')
                r21c3.grid(row=20, column=2)
                r22c3 = Entry(calculation_frame, width=15, justify='center', bg='yellow')
                r22c3.grid(row=21, column=2)
                r23c3 = Entry(calculation_frame, width=15, justify='center', bg='yellow')
                r23c3.grid(row=22, column=2)
                r24c3 = Entry(calculation_frame, width=15, justify='center', bg='yellow')
                r24c3.grid(row=23, column=2)
                r25c3 = Entry(calculation_frame, width=15, justify='center', bg='#96be25')
                r25c3.grid(row=24, column=2)
                quantity_kg = Entry((covert_frame), width=18, justify='center')
                quantity_kg.grid(row=0, column=2)
                quantity_kg_o = Entry((covert_frame), width=18, justify='center')
                quantity_kg_o.grid(row=0, column=0)
                quantity_label_x = Label(covert_frame, text=" X ", width=5)
                quantity_label_x.grid(row=0, column=1)
                quantity_label_kg = Label(covert_frame, text=" kg/m ", width=5)
                quantity_label_kg.grid(row=0, column=3)

                quantity_kg_o.insert(0, 0)
                quantity_kg.insert(0, 0)
                r1c1.insert(0, 0)
                r2c1.insert(0, 0)
                r3c1.insert(0, 0)
                r4c1.insert(0, 0)
                r5c1.insert(0, 0)
                r6c1.insert(0, 0)
                r7c1.insert(0, 0)
                r8c1.insert(0, 0)
                r9c1.insert(0, 0)
                r10c1.insert(0, 0)
                r11c1.insert(0, 0)
                r12c1.insert(0, 0)
                r13c1.insert(0, 0)
                r14c1.insert(0, 0)
                r15c1.insert(0, 0)
                r16c1.insert(0, 0)
                r17c1.insert(0, 0)
                r18c1.insert(0, 0)
                r19c1.insert(0, 0)
                r20c1.insert(0, 0)
                r21c1.insert(0, 0)
                r22c1.insert(0, 0)
                r23c1.insert(0, 0)

                r1c2.insert(0, 0)
                r2c2.insert(0, 0)
                r3c2.insert(0, 0)
                r4c2.insert(0, 0)
                r5c2.insert(0, 0)
                r6c2.insert(0, 0)
                r7c2.insert(0, 0)
                r8c2.insert(0, 0)
                r9c2.insert(0, 0)
                r10c2.insert(0, 0)
                r11c2.insert(0, 0)
                r12c2.insert(0, 0)
                r13c2.insert(0, 0)
                r14c2.insert(0, 0)
                r15c2.insert(0, 0)
                r16c2.insert(0, 0)
                r17c2.insert(0, 0)
                r18c2.insert(0, 0)
                r19c2.insert(0, 0)
                r20c2.insert(0, 0)
                r21c2.insert(0, 0)
                r22c2.insert(0, 0)
                r23c2.insert(0, 0)

                r1c3.insert(0, '')
                r2c3.insert(0, '')
                r3c3.insert(0, '')
                r4c3.insert(0, '')
                r5c3.insert(0, '')
                r6c3.insert(0, '')
                r7c3.insert(0, '')
                r8c3.insert(0, '')
                r9c3.insert(0, '')
                r10c3.insert(0, '')
                r11c3.insert(0, '')
                r12c3.insert(0, '')
                r13c3.insert(0, '')
                r14c3.insert(0, '')
                r15c3.insert(0, '')
                r16c3.insert(0, '')
                r17c3.insert(0, '')
                r18c3.insert(0, '')
                r19c3.insert(0, '')
                r20c3.insert(0, '')
                r21c3.insert(0, '')
                r22c3.insert(0, '')
                r23c3.insert(0, '')
                r24c3.insert(0, '')
                r25c3.insert(0, '')
                rate_box.insert(0, 0)
                quantity_paper_box.insert(0, 0)
                total_cost_box.insert(0, 0)

                # button confirm unit
                confirm = Button(tko, text="CONFIRM UNIT", command=unit_confirm, width=21, bg="yellow")
                confirm.place(x=625, y=420)

                # button calculation
                calculation = Button(tko, text="CALCULATE", width=21, command=calculate)
                calculation.place(x=625, y=450)

                # button back
                calculation = Button(tko, text="BACK", width=21, command=back)
                calculation.place(x=625, y=570)

                # button delete
                calculation = Button(tko, text="CLEAR EXCEL SHEET", width=21, command=delete)
                calculation.place(x=625, y=540)

                # button save to excel
                save_excel = Button(tko, text="SAVE TO EXCEL", command=save, width=21)
                save_excel.place(x=625, y=510)

                # button clear to calculation
                clear_calculation = Button(tko, text="CLEAR CALCULATION", command=clear_calculat, width=21)
                clear_calculation.place(x=625, y=480)

                # canvas
                def locate_xy(event):
                    global current_x, current_y
                    current_x, current_y = event.x, event.y

                def show_color(new_color):
                    global color
                    color = new_color

                def addLine(event):
                    global current_x, current_y
                    if not (event.x < 35 and event.y < 600):
                        canvas.create_line((current_x, current_y, event.x, event.y), width=5, fill=color, capstyle=ROUND,
                                        smooth=True)
                        current_x, current_y = event.x, event.y

                    
                canvas = Canvas(tko, background='white')
                canvas.place(x=304, y=131, height=463, width=306)
                current_x, current_y = 0, 0

                canvas.bind('<Button-1>', locate_xy)
                canvas.bind('<B1-Motion>', addLine)
                canvas.bind("<Double-1>", new_canvas)

                def display_pallete():

                    id = canvas.create_rectangle((10, 10, 30, 30), fill="black")
                    canvas.tag_bind(id, '<Button-1>', lambda x: show_color("black"))

                    # id = canvas.create_rectangle((10, 40, 30, 60), fill="white")
                    # canvas.tag_bind(id, '<Button-1>', lambda x: show_color("white"))

                    id = canvas.create_rectangle((10, 40, 30, 60), fill="brown4")
                    canvas.tag_bind(id, '<Button-1>', lambda x: show_color("brown4"))

                    id = canvas.create_rectangle((10, 70, 30, 90), fill="red")
                    canvas.tag_bind(id, '<Button-1>', lambda x: show_color("red"))

                    id = canvas.create_rectangle((10, 100, 30, 120), fill="yellow")
                    canvas.tag_bind(id, '<Button-1>', lambda x: show_color("yellow"))

                    id = canvas.create_rectangle((10, 130, 30, 150), fill="blue")
                    canvas.tag_bind(id, '<Button-1>', lambda x: show_color('blue'))

                    id = canvas.create_rectangle((10, 160, 30, 180), fill="green")
                    canvas.tag_bind(id, '<Button-1>', lambda x: show_color('green'))

                    id = canvas.create_rectangle((10, 190, 30, 210), fill="purple")
                    canvas.tag_bind(id, '<Button-1>', lambda x: show_color('purple'))

                    id = canvas.create_rectangle((10, 220, 30, 240), fill="orange")
                    canvas.tag_bind(id, '<Button-1>', lambda x: show_color('orange'))
                    
                    # id = canvas.create_rectangle((10, 400, 30, 420), fill="white")
                    # id = canvas.create_text(20,400,fill="black",font="Times 16",text="??")
                    id = canvas.create_text(20,400,fill="black",font="Times 16",text="???")
                    canvas.tag_bind(id, '<Button-1>', lambda x: show_color("white"))
                    
                    id = canvas.create_text(20,450,fill="black",font="Times 16",text="???")
                    canvas.tag_bind(id, '<Button-1>', lambda x: show_color("white"))
                    
                    # issue loading image here
                    # eraserImage = Image.open("eraser.png")
                    # eraserImg = ImageTk.PhotoImage(eraserImage, master=self)
                    # print('eraserImage -> ', eraserImage)
                    # print('eraserImg -> ', eraserImg)
                    # id = canvas.create_image(30, 400, image=eraserImg)
                    # print('id -> ', id)
                    # canvas.tag_bind(id, '<Button-1>', lambda x: show_color("white"))

                display_pallete()

            # remove all
            def remove_all():
                if len(my_tree.get_children()) == 0:
                    messagebox.showerror(message='THERE ARE NO ITEMS TO BE DELETED')
                else:
                    for record in my_tree.get_children():
                        my_tree.delete(record)
                        project_box.delete(0, END)
                        taker_off_name_box.delete(0, END)
                        count = count - 1

            # remove one selected
            def remove_one():
                if len(my_tree.get_children()) == 0:
                    messagebox.showerror(message='THERE ARE NO ITEMS TO BE DELETED')
                elif len(my_tree.selection()) == 0:
                    messagebox.showerror(message='THERE ARE NO ITEMS SELECTED')
                else:
                    x = my_tree.selection()[0]
                    my_tree.delete(x)
                    count = count - 1

            # select record
            def select_record(e):
                # clear entry boxes
                item_box.delete(0, END)
                clause_box.delete(0, END)
                unit_box.delete(0, END)
                quantity_box.delete(0, END)

                # Grab record number
                selected = my_tree.focus()

                # Grab record value
                values = my_tree.item(selected, 'values')
                item_box.insert(0, values[0])
                clause_box.insert(0, values[1])
                unit_box.insert(0, values[2])
                quantity_box.insert(0, values[3])

            # save record
            def save_record():

                if len(my_tree.selection()) == 0:
                    messagebox.showerror(message='THERE ARE NO ITEMS SELECTED')

                else:
                    try:
                        input_quantity = float(quantity_box.get())
                    except ValueError:
                        messagebox.showerror(message="NUMERIC INPUT ONLY")
                        quantity_box.delete(0, END)

                    # Grab record number
                    selected = my_tree.focus()

                    # save new data
                    my_tree.item(selected, text="",
                        values=(item_box.get(), clause_box.get(), unit_box.get(), quantity_box.get()))

                    # clear entry boxes
                    item_box.delete(0, END)
                    clause_box.delete(0, END)
                    unit_box.delete(0, END)
                    quantity_box.delete(0, END)

            def log_out():
                MsgBox = messagebox.askquestion(message='Are you sure to log out ?')
                if MsgBox == 'yes':
                    root.destroy()

            button_frame = Frame(root)
            button_frame.place(x=100, y=440)
            # button add
            add_button = Button(button_frame, text="ADD RECORD", command=add_item, width=15)
            add_button.grid(row=0, column=0, padx=10)

            # button log out
            log_out_button = Button(root, text="LOG OUT", command=log_out, width=15)
            log_out_button.place(x=630, y=550)

            # button taking off paper
            add_taking_off_paper = Button(root, text="TAKING OFF PAPER", command=taking_off_paper)
            add_taking_off_paper.place(x=380, y=475)

            # remove one
            remove_one = Button(button_frame, text="DELETE RECORD", command=remove_one, width=15)
            remove_one.grid(row=0, column=2, padx=10)

            # remove all
            remove_all = Button(button_frame, text="DELETE ALL RECORD", command=remove_all, width=15)
            remove_all.grid(row=0, column=3, padx=10)

            # save update
            save_record = Button(button_frame, text="UPDATE RECORD", command=save_record, width=15)
            save_record.grid(row=0, column=1, padx=10)

            # button query list
            query_list= Button(button_frame, text="QUERY LIST", command= query,width=15)
            query_list.grid(row=2, column=1, pady=10)

            # bind select record
            my_tree.bind("<ButtonRelease-1>", select_record)
            my_tree.bind("<Double-1>", togglecheck)
            root.mainloop()

        else:
            messagebox.showerror(message='Invalid password')

    # PLEASE ENTER THE PASSWORD OR USERNAME
    elif usr_name == '' or usr_pwd == '':
        messagebox.showerror(message='Please enter the username and password.')
    # DATA NOT IN FILE
    else:
        is_signup = messagebox.askyesno('Welcome!', "Account not exist.Do you want to register for FAST Takeoff?")
        if is_signup:
            usr_sign_up()

# SIGN UP FUNCTIOM
def usr_sign_up():
    def signtowcg():
        nn = new_name.get()
        np = new_pwd.get()
        npf = new_pwd_confirm.get()
        try:
            with open('usr_info.pickle', 'rb') as usr_file:
                exist_usr_info = pickle.load(usr_file)
        except FileNotFoundError:
            exist_usr_info = {}

        if nn in exist_usr_info:
            messagebox.showerror("Error!", 'Username already exists. Please try another one.')
        elif np == '' or nn == '':
            messagebox.showerror("Error!", 'Please fill in the username and password.')
        elif np != npf:
            messagebox.showerror("Error!", 'Please re-enter the password.')
        # SAVE THE USER NEW DATA
        else:
            exist_usr_info[nn] = np
            with open('usr_info.pickle', 'wb') as usr_file:
                pickle.dump(exist_usr_info, usr_file)
            messagebox.showinfo("Welcome!", 'Registration success! Please log in.')
            # EXIT THE SIGN UP WINDOW
            window_sign_up.destroy()

    # SIGN UP MAIN WINDOW
    window_sign_up = Toplevel(window)
    window_sign_up.geometry('350x200+480+250')
    window_sign_up.title('SIGN UP')
    # LABEL AND INPUT USERNAME
    new_name = StringVar()
    Label(window_sign_up, text='Username:').place(x=10, y=10)
    Entry(window_sign_up, textvariable=new_name).place(x=160, y=10)
    # LABEL AND INPUT PASSWORD
    new_pwd = StringVar()
    Label(window_sign_up, text='Password:').place(x=10, y=50)
    Entry(window_sign_up, textvariable=new_pwd, show='*').place(x=160, y=50)
    # LABEL AND INPUT REENTER THE PASSWORD
    new_pwd_confirm = StringVar()
    Label(window_sign_up, text='Re-enter the password:').place(x=10, y=90)
    Entry(window_sign_up, textvariable=new_pwd_confirm, show='*').place(x=160, y=90)
    # CONFIRM SIGN UP BUTTON
    bt_confirm_sign_up = Button(window_sign_up, text='CONFIRM SIGN UP',
                                   command=signtowcg)
    bt_confirm_sign_up.place(x=150, y=130)

# EXIT FUNCTION
def usr_sign_quit():
    window.destroy()

# LOG IN, REGISTER AND EXIT BUTTON
bt_login = Button(window, text='LOG IN',bg="#DAF7A6", command=usr_log_in)
bt_login.place(x=100, y=230)
bt_logup = Button(window, text='REGISTER',bg="#DAF7A6", command=usr_sign_up)
bt_logup.place(x=175, y=230)
bt_logquit = Button(window, text='EXIT',bg= "#EE123A", fg= "white", command=usr_sign_quit)
bt_logquit.place(x=260, y=230)

window.mainloop()